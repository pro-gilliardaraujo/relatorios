"""
Script para processamento completo de dados de monitoramento de transbordos.
Lê arquivos TXT ou CSV na pasta raiz, processa-os e gera arquivos Excel com planilhas auxiliares prontas.
Também processa arquivos ZIP contendo TXT ou CSV.
"""

import pandas as pd
import numpy as np
import os
import glob
import zipfile
import tempfile
import shutil
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Configurações
processCsv = True  # Altere para True quando quiser processar arquivos CSV

# Constantes
COLUNAS_REMOVER = [
    'Latitude',
    'Longitude',
    'Regional',
    'Unidade',
    'Centro de Custo',
    'Fazenda', 
    'Zona', 
    'Talhao'
]

COLUNAS_DESEJADAS = [
    'Data', 'Hora', 'Equipamento', 'Descricao Equipamento', 'Estado', 'Estado Operacional',
    'Grupo Equipamento/Frente', 'Grupo Operacao', 'Horimetro', 'Motor Ligado', 'Operacao', 'Operador',
    'RPM Motor', 'Tipo de Equipamento', 'Velocidade', 'Parado com motor ligado',
    'Diferença_Hora', 'Horas Produtivas', 'GPS'
]

# Valores a serem filtrados
OPERADORES_EXCLUIR = ["9999 - TROCA DE TURNO"]

def carregar_config_calculos():
    """
    Carrega as configurações de cálculos do arquivo JSON.
    Se o arquivo não existir, retorna configurações padrão.
    """
    config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config", "calculos_config.json")
    
    # Configuração padrão
    config_padrao = {
        "CD": {
            "motor_ocioso": {
                "tipo_calculo": "Remover do cálculo",
                "operacoes_excluidas": [],
                "grupos_operacao_excluidos": [],
                "operadores_excluidos": []
            },
            "equipamentos_excluidos": []
        },
        "TT": {
            "motor_ocioso": {
                "tipo_calculo": "Remover do cálculo",
                "operacoes_excluidas": [],
                "grupos_operacao_excluidos": [],
                "operadores_excluidos": []
            },
            "equipamentos_excluidos": []
        }
    }
    
    try:
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
                
                # Garantir que os equipamentos excluídos sejam tratados como texto
                for tipo in ["CD", "TT"]:
                    if tipo in config and "equipamentos_excluidos" in config[tipo]:
                        config[tipo]["equipamentos_excluidos"] = [str(eq).replace('.0', '') for eq in config[tipo]["equipamentos_excluidos"]]
                
                return config
        else:
            # Criar diretório config se não existir
            config_dir = os.path.dirname(config_path)
            if not os.path.exists(config_dir):
                os.makedirs(config_dir)
                
            # Criar arquivo de configuração padrão
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config_padrao, f, indent=4, ensure_ascii=False)
                
            print(f"Arquivo de configuração criado em {config_path} com valores padrão.")
            return config_padrao
    except Exception as e:
        print(f"Erro ao carregar configurações: {str(e)}. Usando configuração padrão.")
        return config_padrao

def carregar_substituicoes_operadores():
    """
    Carrega o arquivo substituiroperadores.json que contém os mapeamentos 
    de substituição de operadores.
    
    Returns:
        dict: Dicionário com mapeamento {operador_origem: operador_destino}
        ou dicionário vazio se o arquivo não existir ou for inválido
    """
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    
    # Caminho para o arquivo de substituição
    arquivo_substituicao = os.path.join(diretorio_raiz, "config", "substituiroperadores.json")
    
    # Verificar se o arquivo existe
    if not os.path.exists(arquivo_substituicao):
        print(f"Arquivo de substituição de operadores não encontrado: {arquivo_substituicao}")
        return {}
    
    try:
        # Carregar o arquivo JSON
        with open(arquivo_substituicao, 'r', encoding='utf-8') as f:
            substituicoes = json.load(f)
        
        # Criar dicionário de substituições
        mapeamento = {item['operador_origem']: item['operador_destino'] for item in substituicoes}
        
        print(f"Carregadas {len(mapeamento)} substituições de operadores.")
        return mapeamento
        
    except Exception as e:
        print(f"Erro ao carregar arquivo de substituição de operadores: {str(e)}")
        return {}

def aplicar_substituicao_operadores(df, mapeamento_substituicoes):
    """
    Aplica as substituições de operadores no DataFrame.
    
    Args:
        df (DataFrame): DataFrame a ser processado
        mapeamento_substituicoes (dict): Dicionário com mapeamento {operador_origem: operador_destino}
    
    Returns:
        tuple: (DataFrame com substituições aplicadas, DataFrame com registro das substituições)
    """
    if not mapeamento_substituicoes or 'Operador' not in df.columns:
        return df, pd.DataFrame(columns=['ID Original', 'Nome Original', 'ID Nova', 'Nome Novo', 'Registros Afetados'])
    
    # Criar uma cópia para não alterar o DataFrame original
    df_modificado = df.copy()
    
    # Lista para armazenar as substituições realizadas
    substituicoes_realizadas = []
    
    # Contar operadores antes da substituição
    contagem_antes = df_modificado['Operador'].value_counts()
    
    # Aplicar as substituições
    df_modificado['Operador'] = df_modificado['Operador'].replace(mapeamento_substituicoes)
    
    # Contar operadores depois da substituição
    contagem_depois = df_modificado['Operador'].value_counts()
    
    # Verificar quais operadores foram substituídos
    for operador_origem, operador_destino in mapeamento_substituicoes.items():
        if operador_origem in contagem_antes:
            registros_afetados = contagem_antes.get(operador_origem, 0)
            if registros_afetados > 0:
                # Extrair IDs e nomes
                id_original = operador_origem.split(' - ')[0] if ' - ' in operador_origem else operador_origem
                nome_original = operador_origem.split(' - ')[1] if ' - ' in operador_origem else ''
                id_nova = operador_destino.split(' - ')[0] if ' - ' in operador_destino else operador_destino
                nome_novo = operador_destino.split(' - ')[1] if ' - ' in operador_destino else ''
                
                substituicoes_realizadas.append({
                    'ID Original': id_original,
                    'Nome Original': nome_original,
                    'ID Nova': id_nova,
                    'Nome Novo': nome_novo,
                    'Registros Afetados': registros_afetados
                })
                print(f"Operador '{operador_origem}' substituído por '{operador_destino}' em {registros_afetados} registros")
    
    # Criar DataFrame com as substituições realizadas
    df_substituicoes = pd.DataFrame(substituicoes_realizadas)
    
    return df_modificado, df_substituicoes

def processar_arquivo_base(caminho_arquivo):
    """
    Processa o arquivo TXT ou CSV de transbordos e retorna o DataFrame com as transformações necessárias.
    Usando exatamente o mesmo método do Codigo_Base_TT.py para cálculo da Diferença_Hora.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo TXT ou CSV de entrada
    
    Returns:
        DataFrame: DataFrame processado com todas as transformações
    """
    # Lista de codificações para tentar
    codificacoes = ['utf-8', 'latin1', 'ISO-8859-1', 'cp1252']
    
    for codificacao in codificacoes:
        try:
            # Leitura do arquivo (TXT ou CSV são tratados da mesma forma se usarem separador ';')
            df = pd.read_csv(caminho_arquivo, sep=';', encoding=codificacao)
            print(f"Arquivo lido com sucesso usando {codificacao}! Total de linhas: {len(df)}")
            
            # Verificar se o DataFrame está vazio (apenas cabeçalhos sem dados)
            if len(df) == 0:
                print(f"O arquivo {caminho_arquivo} contém apenas cabeçalhos sem dados.")
                # Retornar o DataFrame vazio mas com as colunas, em vez de None
                # Garantir que todas as colunas desejadas existam
                for col in COLUNAS_DESEJADAS:
                    if col not in df.columns:
                        df[col] = np.nan
                # Reorganizar as colunas na ordem desejada
                colunas_existentes = [col for col in COLUNAS_DESEJADAS if col in df.columns]
                colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS]
                return df[colunas_existentes + colunas_extras]
            
            # Limpeza de espaços extras nos nomes das colunas
            df.columns = df.columns.str.strip()
            
            # Padronizar valores da coluna Grupo Operacao
            if 'Grupo Operacao' in df.columns:
                df['Grupo Operacao'] = df['Grupo Operacao'].str.strip()
                # Mapear valores para garantir consistência
                mapa_grupo_operacao = {
                    'SEM APONTAMENTO': 'Sem Apontamento',
                    'PRODUTIVA': 'Produtiva',
                    'MANUTENCAO': 'Manutenção',
                    'MANUTENÇÃO': 'Manutenção'
                }
                df['Grupo Operacao'] = df['Grupo Operacao'].replace(mapa_grupo_operacao)
            
            # Verificar se 'Data/Hora' existe, caso ainda não tenha sido separado
            if 'Data/Hora' in df.columns:
                df[['Data', 'Hora']] = df['Data/Hora'].str.split(' ', expand=True)
                df = df.drop(columns=['Data/Hora'])
            
            # Remover colunas conforme solicitado no Codigo_Base_TT.py
            colunas_remover = ['Unidade', 'Centro de Custo', 'Fazenda', 'Zona', 'Talhao']
            df = df.drop(columns=colunas_remover, errors='ignore')
            
            # MÉTODO EXATO DO Codigo_Base_TT.py para cálculo da Diferença_Hora para garantir mesmos resultados
            # Conversão de Hora para datetime (apenas se ainda não for)
            if df['Hora'].dtype != 'datetime64[ns]':
                df['Hora'] = pd.to_datetime(df['Hora'], format='%H:%M:%S', errors='coerce')
            
            # Calcular Diferença_Hora sem arredondamentos usando o método EXATO do Codigo_Base_TT.py
            # NOTA: Removida a regra que zerava valores > 0.5, pois não existe no Codigo_Base_TT.py
            # e estava causando perda de aproximadamente 16 horas no total
            df['Diferença_Hora'] = pd.to_datetime(df['Hora'], format='%H:%M:%S').diff()
            df['Diferença_Hora'] = pd.to_timedelta(df['Diferença_Hora'], errors='coerce')
            df['Diferença_Hora'] = df['Diferença_Hora'].dt.total_seconds() / 3600  # Conversor para horas
            df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: x if x >= 0 else 0)
            
            # Soma total para verificação de precisão (mesma lógica do Codigo_Base_TT.py)
            print(f"Diferença_Hora calculada usando método exato do Codigo_Base_TT.py. Soma total: {df['Diferença_Hora'].sum():.8f} horas")
            
            # Conversão de Motor Ligado para formato conforme Codigo_Base_TT.py
            for col in ['Motor Ligado']:
                if col in df.columns:
                    df[col] = df[col].replace({1: 'LIGADO', 0: 'DESLIGADO'})
            
            # Criar a coluna "Parado com motor ligado" exatamente como no Codigo_Base_TT.py
            df['Parado com motor ligado'] = ((df['Velocidade'] == 0) & (df['Motor Ligado'] == 'LIGADO')).astype(int)
            
            # Verifica se Horas Produtivas já existe, senão calcula usando método do Codigo_Base_TT.py
            if 'Horas Produtivas' not in df.columns or df['Horas Produtivas'].isna().any():
                # Calcular horas produtivas sem arredondamento, mantendo a precisão completa
                df['Horas Produtivas'] = df.apply(
                    lambda row: row['Diferença_Hora'] if row['Grupo Operacao'] == 'Produtiva' else 0,
                    axis=1
                )
                # Soma total de horas produtivas para verificação
                print(f"Total de horas produtivas: {df['Horas Produtivas'].sum():.8f}")
            else:
                # Limpa e converte para número
                df['Horas Produtivas'] = pd.to_numeric(df['Horas Produtivas'].astype(str).str.strip(), errors='coerce')
                df['Horas Produtivas'] = df['Horas Produtivas'].fillna(0)
            
            # Coluna de GPS - Para transbordos, vamos considerar GPS quando houver "RTK (Piloto Automatico)" 
            # e Velocidade > 0 (se a coluna existir)
            if 'RTK (Piloto Automatico)' in df.columns:
                df['GPS'] = df.apply(
                    lambda row: row['Diferença_Hora'] if row.get('RTK (Piloto Automatico)', 0) == 1 
                    and row['Velocidade'] > 0 and row['Grupo Operacao'] == 'Produtiva' else 0,
                    axis=1
                )
            else:
                # Se não tiver a coluna RTK, criar uma coluna GPS zerada
                df['GPS'] = 0
            
            # Conversão de colunas binárias para valores numéricos (garantindo que sejam números)
            for col in ['Esteira Ligada', 'Field Cruiser', 'RTK (Piloto Automatico)', 'Implemento Ligado']:
                if col in df.columns and col != 'Motor Ligado':  # Motor Ligado já foi tratado acima
                    # Se a coluna for texto (LIGADO/DESLIGADO), converter para 1/0
                    if df[col].dtype == 'object':
                        df[col] = df[col].replace({'LIGADO': 1, 'DESLIGADO': 0})
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
            
            # Limpeza e organização das colunas
            df = df.drop(columns=COLUNAS_REMOVER, errors='ignore')
            
            # Garantir que todas as colunas desejadas existam
            for col in COLUNAS_DESEJADAS:
                if col not in df.columns:
                    df[col] = np.nan
            
            # Reorganizar as colunas na ordem desejada
            colunas_existentes = [col for col in COLUNAS_DESEJADAS if col in df.columns]
            colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS]
            df = df[colunas_existentes + colunas_extras]
            
            return df
            
        except UnicodeDecodeError:
            print(f"Tentativa com codificação {codificacao} falhou, tentando próxima codificação...")
            continue
        except Exception as e:
            print(f"Erro ao processar o arquivo com codificação {codificacao}: {str(e)}")
            continue
    
    # Se chegou aqui, todas as tentativas de codificação falharam
    print(f"Erro: Não foi possível ler o arquivo {caminho_arquivo} com nenhuma das codificações tentadas.")
    return None

def calcular_base_calculo(df):
    """
    Calcula a tabela de Base Calculo a partir do DataFrame processado.
    Calcula médias diárias considerando os dias efetivos de trabalho de cada operador.
    
    Cálculos principais:
    - Horas totais: soma de Diferença_Hora
    - Motor Ligado: soma de Diferença_Hora onde Motor Ligado = LIGADO
    - Parado com motor ligado: soma de Diferença_Hora onde Motor Ligado = LIGADO E Velocidade = 0
    - GPS: soma de Diferença_Hora onde RTK = 1
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Tabela Base Calculo com todas as métricas calculadas
    """
    # Detectar número de dias totais nos dados (apenas para informação)
    dias_unicos_total = df['Data'].nunique() if 'Data' in df.columns else 1
    print(f"Detectados {dias_unicos_total} dias distintos na base de dados.")
    
    # Extrair combinações únicas de Equipamento, Grupo Equipamento/Frente e Operador
    combinacoes = df[['Equipamento', 'Grupo Equipamento/Frente', 'Operador']].drop_duplicates().reset_index(drop=True)
    
    # Filtrar operadores excluídos
    combinacoes = combinacoes[~combinacoes['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Inicializar as colunas de métricas
    resultados = []
    
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    # Calcular as métricas para cada combinação
    for idx, row in combinacoes.iterrows():
        equipamento = row['Equipamento']
        grupo = row['Grupo Equipamento/Frente']
        operador = row['Operador']
        
        # Filtrar dados para esta combinação
        filtro = (df['Equipamento'] == equipamento) & \
                (df['Grupo Equipamento/Frente'] == grupo) & \
                (df['Operador'] == operador)
        
        dados_filtrados = df[filtro]
        
        # Determinar o número de dias efetivos para este operador
        dias_operador = dados_filtrados['Data'].nunique() if 'Data' in dados_filtrados.columns else 1
        
        # Horas totais - soma de Diferença_Hora
        horas_totais = dados_filtrados['Diferença_Hora'].sum()
        if dias_operador > 1:
            horas_totais = horas_totais / dias_operador
        
        # Motor Ligado - soma de Diferença_Hora onde Motor Ligado = LIGADO
        motor_ligado = dados_filtrados[
            dados_filtrados['Motor Ligado'] == 'LIGADO'
        ]['Diferença_Hora'].sum()
        if dias_operador > 1:
            motor_ligado = motor_ligado / dias_operador
        
        # Parado com motor ligado - soma de Diferença_Hora onde Motor Ligado = LIGADO E Velocidade = 0
        parado_motor_ligado = dados_filtrados[
            (dados_filtrados['Motor Ligado'] == 'LIGADO') & 
            (dados_filtrados['Velocidade'] == 0)
        ]['Diferença_Hora'].sum()
        if dias_operador > 1:
            parado_motor_ligado = parado_motor_ligado / dias_operador
        
        # % Parado com motor ligado (em decimal 0-1)
        percent_parado_motor = calcular_porcentagem(parado_motor_ligado, motor_ligado)
        
        # GPS - soma de Diferença_Hora onde RTK = 1
        gps = dados_filtrados[
            dados_filtrados['RTK (Piloto Automatico)'] == 1
        ]['Diferença_Hora'].sum()
        if dias_operador > 1:
            gps = gps / dias_operador
        
        # Horas Produtivas
        horas_produtivas = dados_filtrados[
            dados_filtrados['Grupo Operacao'] == 'Produtiva'
        ]['Diferença_Hora'].sum()
        if dias_operador > 1:
            horas_produtivas = horas_produtivas / dias_operador
        
        # % Utilização GPS (em decimal 0-1)
        utilizacao_gps = calcular_porcentagem(gps, horas_produtivas)
        
        # Falta de Apontamento - soma de Diferença_Hora onde Operacao = '8340 - FALTA DE APONTAMENTO'
        falta_apontamento = dados_filtrados[
            dados_filtrados['Operacao'] == '8340 - FALTA DE APONTAMENTO'
        ]['Diferença_Hora'].sum()
        if dias_operador > 1:
            falta_apontamento = falta_apontamento / dias_operador
        
        # % Falta de Apontamento (em decimal 0-1)
        percent_falta_apontamento = calcular_porcentagem(falta_apontamento, motor_ligado)
        
        # Debug para verificar os valores
        print(f"\nOperador: {operador} em {equipamento}")
        print(f"Motor Ligado: {motor_ligado:.6f}")
        print(f"Parado com motor ligado: {parado_motor_ligado:.6f}")
        print(f"% Parado com motor ligado: {percent_parado_motor:.6f}")
        print(f"Falta de Apontamento: {falta_apontamento:.6f}")
        print(f"% Falta de Apontamento: {percent_falta_apontamento:.6f}")
        
        resultados.append({
            'Equipamento': equipamento,
            'Grupo Equipamento/Frente': grupo,
            'Operador': operador,
            'Horas totais': horas_totais,
            'Motor Ligado': motor_ligado,
            'Parado com motor ligado': parado_motor_ligado,
            '% Parado com motor ligado': percent_parado_motor,
            'GPS': gps,
            'Horas Produtivas': horas_produtivas,
            '% Utilização GPS': utilizacao_gps,
            'Falta de Apontamento': falta_apontamento,
            '% Falta de Apontamento': percent_falta_apontamento
        })
    
    return pd.DataFrame(resultados)

def calcular_disponibilidade_mecanica(df):
    """
    Calcula a disponibilidade mecânica para cada equipamento.
    Calcula médias diárias considerando os dias efetivos de cada equipamento.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Disponibilidade mecânica por equipamento
    """
    # Filtramos os dados excluindo os operadores da lista
    df_filtrado = df[~df['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    # Agrupar por Equipamento e calcular horas por grupo operacional
    equipamentos = df_filtrado['Equipamento'].unique()
    resultados = []
    
    for equipamento in equipamentos:
        dados_equip = df_filtrado[df_filtrado['Equipamento'] == equipamento]
        
        # Determinar número de dias efetivos para este equipamento
        dias_equip = dados_equip['Data'].nunique() if 'Data' in dados_equip.columns else 1
        
        total_horas = dados_equip['Diferença_Hora'].sum()
        
        # Calcular horas de manutenção
        manutencao = dados_equip[dados_equip['Grupo Operacao'] == 'Manutenção']['Diferença_Hora'].sum()
        
        # Se houver múltiplos dias, usar médias diárias
        if dias_equip > 1:
            total_horas = total_horas / dias_equip
            manutencao = manutencao / dias_equip
            print(f"Equipamento: {equipamento}, Dias efetivos: {dias_equip}, Média diária: {total_horas:.6f} horas")
        
        # A disponibilidade mecânica é o percentual de tempo fora de manutenção
        disp_mecanica = calcular_porcentagem(total_horas - manutencao, total_horas)
        
        resultados.append({
            'Frota': equipamento,
            'Disponibilidade': disp_mecanica
        })
    
    return pd.DataFrame(resultados)

def calcular_horas_por_frota(df):
    """
    Calcula o total de horas registradas para cada frota e a diferença para 24 horas.
    Calcula médias diárias considerando os dias efetivos de cada frota.
    Esta função NÃO aplica qualquer filtro de operador.
    Também identifica as faltas de horário por dia específico.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Horas totais por frota com detalhamento por dia
    """
    # Agrupar por Equipamento e somar as diferenças de hora
    equipamentos = df['Equipamento'].unique()
    resultados = []
    
    # Obter todos os dias únicos no dataset
    dias_unicos = sorted(df['Data'].unique()) if 'Data' in df.columns else []
    
    for equipamento in equipamentos:
        dados_equip = df[df['Equipamento'] == equipamento]
        
        # Determinar número de dias efetivos para este equipamento
        dias_equip = dados_equip['Data'].nunique() if 'Data' in dados_equip.columns else 1
        
        total_horas = dados_equip['Diferença_Hora'].sum()
        
        # Se houver múltiplos dias, usar média diária
        if dias_equip > 1:
            total_horas = total_horas / dias_equip
        
        # Calcular a diferença para 24 horas
        diferenca_24h = max(24 - total_horas, 0)
        
        # Criar o resultado básico (colunas originais mantidas)
        resultado = {
            'Frota': equipamento,
            'Horas Registradas': total_horas,
            'Diferença para 24h': diferenca_24h
        }
        
        # Adicionar detalhamento por dia (novas colunas)
        if len(dias_unicos) > 0:
            for dia in dias_unicos:
                dados_dia = dados_equip[dados_equip['Data'] == dia]
                
                # Se não houver dados para este dia e equipamento, a diferença é 24h
                if len(dados_dia) == 0:
                    resultado[f'Falta {dia}'] = 24.0
                    continue
                
                # Calcular horas registradas neste dia
                horas_dia = dados_dia['Diferença_Hora'].sum()
                
                # Calcular a diferença para 24 horas neste dia
                diferenca_dia = max(24 - horas_dia, 0)
                
                # Adicionar ao resultado apenas se houver falta (diferença > 0)
                if diferenca_dia > 0:
                    resultado[f'Falta {dia}'] = diferenca_dia
                else:
                    resultado[f'Falta {dia}'] = 0.0
        
        resultados.append(resultado)
    
    return pd.DataFrame(resultados)

def calcular_eficiencia_energetica(base_calculo):
    """
    Calcula a eficiência energética por operador usando os dados da Base Calculo.
    Agora usa diretamente os valores calculados na Base Calculo ao invés de recalcular.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Eficiência energética por operador (agregado)
    """
    # Agrupar por operador e calcular a média ponderada
    agrupado = base_calculo.groupby('Operador').agg({
        'Motor Ligado': 'sum',
        'Horas Produtivas': 'sum'
    }).reset_index()
    
    resultados = []
    for _, row in agrupado.iterrows():
        # Calcular eficiência como Horas Produtivas / Motor Ligado
        eficiencia = row['Horas Produtivas'] / row['Motor Ligado'] if row['Motor Ligado'] > 0 else 0
        
        resultados.append({
            'Operador': row['Operador'],
            'Eficiência': eficiencia
        })
    
    return pd.DataFrame(resultados)

def calcular_motor_ocioso(base_calculo, df_base):
    """
    Calcula o percentual de motor ocioso por operador usando os dados da Base Calculo.
    Agora usa diretamente os valores calculados na Base Calculo ao invés de recalcular.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
        df_base (DataFrame): DataFrame base (não usado mais, mantido para compatibilidade)
    
    Returns:
        DataFrame: Percentual de motor ocioso por operador com as colunas:
            - Operador
            - Porcentagem
            - Tempo Ligado (vem da coluna 'Motor Ligado' da Base Calculo)
            - Tempo Ocioso (vem da coluna 'Parado com motor ligado' da Base Calculo)
    """
    # Agrupar por operador (caso o mesmo operador apareça em múltiplas linhas)
    agrupado = base_calculo.groupby('Operador').agg({
        'Motor Ligado': 'sum',
        'Parado com motor ligado': 'sum'  # Nome correto da coluna para transbordos
    }).reset_index()
    
    resultados = []
    print("\n=== DETALHAMENTO DO CÁLCULO DE MOTOR OCIOSO (USANDO BASE CALCULO) ===")
    
    for _, row in agrupado.iterrows():
        tempo_ligado = row['Motor Ligado']
        tempo_ocioso = row['Parado com motor ligado']  # Nome correto da coluna para transbordos
        
        # Calcular porcentagem
        porcentagem = tempo_ocioso / tempo_ligado if tempo_ligado > 0 else 0
        
        print(f"\nOperador: {row['Operador']}")
        print(f"Tempo Ligado (Motor Ligado): {tempo_ligado:.6f}")
        print(f"Tempo Ocioso (Parado com motor ligado): {tempo_ocioso:.6f}")
        print(f"Porcentagem: {porcentagem:.6f}")
        print("-" * 50)
        
        resultados.append({
            'Operador': row['Operador'],
            'Porcentagem': porcentagem,
            'Tempo Ligado': tempo_ligado,
            'Tempo Ocioso': tempo_ocioso
        })
    
    return pd.DataFrame(resultados)

def calcular_falta_apontamento(base_calculo):
    """
    Calcula a falta de apontamento por operador usando os dados da Base Calculo.
    Agora usa diretamente os valores calculados na Base Calculo ao invés de recalcular.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Falta de apontamento por operador (agregado)
    """
    # Agrupar por operador e calcular a média ponderada
    agrupado = base_calculo.groupby('Operador').agg({
        'Motor Ligado': 'sum',
        'Falta de Apontamento': 'sum'
    }).reset_index()
    
    resultados = []
    for _, row in agrupado.iterrows():
        # Calcular porcentagem de falta de apontamento
        porcentagem = row['Falta de Apontamento'] / row['Motor Ligado'] if row['Motor Ligado'] > 0 else 0
        
        resultados.append({
            'Operador': row['Operador'],
            'Porcentagem': porcentagem
        })
    
    return pd.DataFrame(resultados)

def calcular_uso_gps(base_calculo):
    """
    Calcula o uso de GPS por operador usando os dados da Base Calculo.
    Agora usa diretamente os valores calculados na Base Calculo ao invés de recalcular.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Uso de GPS por operador (agregado)
    """
    # Agrupar por operador e calcular a média ponderada
    agrupado = base_calculo.groupby('Operador').agg({
        'GPS': 'sum',
        'Horas Produtivas': 'sum'
    }).reset_index()
    
    resultados = []
    for _, row in agrupado.iterrows():
        # Calcular porcentagem de uso de GPS
        porcentagem = row['GPS'] / row['Horas Produtivas'] if row['Horas Produtivas'] > 0 else 0
        
        resultados.append({
            'Operador': row['Operador'],
            'Porcentagem': porcentagem
        })
    
    return pd.DataFrame(resultados)

def calcular_media_velocidade(df):
    """
    Calcula a média de velocidade para cada operador.
    
    Args:
        df (DataFrame): DataFrame com os dados
        
    Returns:
        DataFrame: DataFrame com a média de velocidade por operador
    """
    # Filtrar operadores excluídos
    df = df[~df['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Identificar registros válidos para cálculo de velocidade
    # Usar 'Grupo Operacao' == 'Produtiva' em vez de 'Produtivo' == 1
    registros_validos = (df['Grupo Operacao'] == 'Produtiva') & (df['Velocidade'] > 0)
    
    # Se a coluna 'Movimento' existir, adicionar à condição
    if 'Movimento' in df.columns:
        registros_validos = registros_validos & (df['Movimento'] == 1)
    
    # Calcular média de velocidade por operador
    media_velocidade = df[registros_validos].groupby('Operador')['Velocidade'].mean().reset_index()
    
    # Garantir que todos os operadores estejam no resultado, mesmo sem velocidade
    todos_operadores = df['Operador'].unique()
    for operador in todos_operadores:
        if operador not in media_velocidade['Operador'].values:
            media_velocidade = pd.concat([
                media_velocidade,
                pd.DataFrame({'Operador': [operador], 'Velocidade': [0]})
            ], ignore_index=True)
    
    # Ordenar por operador
    media_velocidade = media_velocidade.sort_values('Operador')
    
    return media_velocidade

def identificar_operadores_duplicados(df, substituicoes=None):
    """
    Identifica operadores que começam com '133' e têm 7 dígitos.
    Verifica se já existe uma substituição no arquivo JSON, caso contrário, registra como ID encontrada.
    
    Args:
        df (DataFrame): DataFrame com os dados dos operadores
        substituicoes (dict): Dicionário com as substituições do arquivo JSON
    
    Returns:
        dict: Dicionário com mapeamento {id_incorreta: id_correta}
        DataFrame: DataFrame com as IDs encontradas para relatório
    """
    if 'Operador' not in df.columns or len(df) == 0:
        return {}, pd.DataFrame(columns=['ID Encontrada', 'Nome', 'Status', 'ID Substituição'])
    
    # Extrair operadores únicos
    operadores = df['Operador'].unique()
    
    # Lista para armazenar as IDs encontradas
    ids_encontradas = []
    mapeamento = {}
    
    for op in operadores:
        if ' - ' in op:
            try:
                id_parte, nome_parte = op.split(' - ', 1)
                # Verificar se a ID começa com 133 e tem 7 dígitos
                if id_parte.startswith('133') and len(id_parte) == 7:
                    # Verificar se existe uma substituição no arquivo JSON
                    if substituicoes and op in substituicoes:
                        status = "Substituição encontrada"
                        id_substituicao = substituicoes[op].split(' - ')[0] if ' - ' in substituicoes[op] else substituicoes[op]
                        mapeamento[op] = substituicoes[op]
                    else:
                        status = "Sem substituição definida"
                        id_substituicao = ""
                    
                    # Adicionar à lista de IDs encontradas, mesmo se for "NAO CADASTRADO"
                    ids_encontradas.append({
                        'ID Encontrada': id_parte,
                        'Nome': nome_parte,
                        'Status': status,
                        'ID Substituição': id_substituicao
                    })
                    
                    print(f"ID encontrada: {id_parte} - {nome_parte}")
            except Exception as e:
                print(f"Erro ao processar operador {op}: {str(e)}")
                continue
    
    print(f"Encontradas {len(ids_encontradas)} IDs começando com 133 e 7 dígitos.")
    for id_enc in ids_encontradas:
        print(f"  - {id_enc['ID Encontrada']} - {id_enc['Nome']} ({id_enc['Status']})")
    
    # Criar o DataFrame e ordenar por ID Encontrada
    df_encontradas = pd.DataFrame(ids_encontradas)
    if not df_encontradas.empty:
        df_encontradas = df_encontradas.sort_values('ID Encontrada')
    
    return mapeamento, df_encontradas

def criar_excel_com_planilhas(df_base, base_calculo, disp_mecanica, eficiencia_energetica,
                            motor_ocioso, falta_apontamento, uso_gps, horas_por_frota, caminho_saida,
                            df_duplicados=None, media_velocidade=None, df_substituicoes=None):
    """
    Cria um arquivo Excel com todas as planilhas necessárias.
    """
    # Definir função de ajuste de largura de colunas
    def ajustar_largura_colunas(worksheet):
        """Ajusta a largura das colunas da planilha"""
        for col in worksheet.columns:
            max_length = 10
            column = col[0].column_letter
            header_text = str(col[0].value)
            if header_text:
                max_length = max(max_length, len(header_text) + 2)
            for cell in col[1:min(20, len(col))]:
                if cell.value:
                    cell_text = str(cell.value)
                    max_length = max(max_length, len(cell_text) + 2)
            max_length = min(max_length, 40)
            worksheet.column_dimensions[column].width = max_length
    
    with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
        # Salvar cada DataFrame em uma planilha separada
        df_base.to_excel(writer, sheet_name='BASE', index=False)
        base_calculo.to_excel(writer, sheet_name='Base Calculo', index=False)
        disp_mecanica.to_excel(writer, sheet_name='1_Disponibilidade Mecânica', index=False)
        eficiencia_energetica.to_excel(writer, sheet_name='2_Eficiência Energética', index=False)
        
        # Garantir que os valores numéricos do motor_ocioso sejam mantidos como números
        motor_ocioso['Tempo Ligado'] = pd.to_numeric(motor_ocioso['Tempo Ligado'], errors='coerce')
        motor_ocioso['Tempo Ocioso'] = pd.to_numeric(motor_ocioso['Tempo Ocioso'], errors='coerce')
        motor_ocioso['Porcentagem'] = pd.to_numeric(motor_ocioso['Porcentagem'], errors='coerce')
        motor_ocioso.to_excel(writer, sheet_name='3_Motor Ocioso', index=False)
        
        falta_apontamento.to_excel(writer, sheet_name='4_Falta Apontamento', index=False)
        uso_gps.to_excel(writer, sheet_name='5_Uso GPS', index=False)
        horas_por_frota.to_excel(writer, sheet_name='Horas por Frota', index=False)
        
        if media_velocidade is None:
            media_velocidade = pd.DataFrame(columns=['Operador', 'Velocidade'])
        media_velocidade.to_excel(writer, sheet_name='Média Velocidade', index=False)
        
        # IDs duplicadas e substituídas
        if df_duplicados is not None and not df_duplicados.empty:
            df_duplicados.to_excel(writer, sheet_name='IDs Encontradas', index=False)
        if df_substituicoes is not None and not df_substituicoes.empty:
            df_substituicoes.to_excel(writer, sheet_name='IDs Substituídas', index=False)
        
        # Formatar cada planilha
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            ajustar_largura_colunas(worksheet)
            
            if sheet_name == '1_Disponibilidade Mecânica':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Disponibilidade)
                    cell.number_format = '0.00%'
            
            elif sheet_name == '2_Eficiência Energética':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Eficiência)
                    cell.number_format = '0.00%'
            
            elif sheet_name == '3_Motor Ocioso':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
                    cell.number_format = '0.00%'
                    cell = worksheet.cell(row=row, column=3)  # Coluna C (Tempo Ligado)
                    cell.number_format = '0.00'
                    cell = worksheet.cell(row=row, column=4)  # Coluna D (Tempo Ocioso)
                    cell.number_format = '0.00'
            
            elif sheet_name == '4_Falta Apontamento':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
                    cell.number_format = '0.00%'
            
            elif sheet_name == '5_Uso GPS':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
                    cell.number_format = '0.00%'
            
            elif sheet_name == 'Média Velocidade':
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=2)  # Coluna B (Velocidade)
                    cell.number_format = '0.00'
            
            elif sheet_name == 'Horas por Frota':
                for row in range(2, worksheet.max_row + 1):
                    for col in range(2, worksheet.max_column + 1):  # Todas as colunas de tempo
                        cell = worksheet.cell(row=row, column=col)
                        cell.number_format = '0.00'
            
            elif sheet_name == 'Base Calculo':
                colunas_porcentagem = ['% Parado com motor ligado', '% Utilização GPS', '% Falta de Apontamento']
                colunas_tempo = ['Horas totais', 'Motor Ligado', 'Parado com motor ligado', 'GPS', 'Horas Produtivas', 'Falta de Apontamento']
                
                for row in range(2, worksheet.max_row + 1):
                    for col in range(1, worksheet.max_column + 1):
                        header = worksheet.cell(row=1, column=col).value
                        cell = worksheet.cell(row=row, column=col)
                        
                        if header in colunas_porcentagem:
                            cell.number_format = '0.00%'
                        elif header in colunas_tempo:
                            cell.number_format = '0.00'

def extrair_arquivo_zip(caminho_zip, pasta_destino=None):
    """
    Extrai o conteúdo de um arquivo ZIP para uma pasta temporária ou destino especificado.
    Renomeia os arquivos extraídos para terem o mesmo nome do arquivo ZIP original.
    
    Args:
        caminho_zip (str): Caminho para o arquivo ZIP
        pasta_destino (str, optional): Pasta onde os arquivos serão extraídos.
                                       Se None, usa uma pasta temporária.
    
    Returns:
        list: Lista de caminhos dos arquivos extraídos e renomeados (apenas TXT e CSV)
        str: Caminho da pasta temporária (se criada) ou None
    """
    # Se pasta_destino não foi especificada, criar uma pasta temporária
    pasta_temp = None
    if pasta_destino is None:
        pasta_temp = tempfile.mkdtemp()
        pasta_destino = pasta_temp
    
    arquivos_extraidos = []
    nome_zip_sem_extensao = os.path.splitext(os.path.basename(caminho_zip))[0]
    
    try:
        with zipfile.ZipFile(caminho_zip, 'r') as zip_ref:
            # Extrair todos os arquivos do ZIP
            zip_ref.extractall(pasta_destino)
            
            # Processar cada arquivo extraído (apenas TXT e CSV)
            for arquivo in zip_ref.namelist():
                caminho_completo = os.path.join(pasta_destino, arquivo)
                # Verificar se é um arquivo e não uma pasta
                if os.path.isfile(caminho_completo):
                    # Verificar extensão
                    extensao = os.path.splitext(arquivo)[1].lower()
                    if extensao in ['.txt', '.csv']:
                        # Criar novo nome: nome do ZIP + extensão original
                        novo_nome = f"{nome_zip_sem_extensao}{extensao}"
                        novo_caminho = os.path.join(pasta_destino, novo_nome)
                        
                        # Renomear o arquivo extraído
                        try:
                            # Se já existe um arquivo com esse nome, remover primeiro
                            if os.path.exists(novo_caminho):
                                os.remove(novo_caminho)
                            # Renomear o arquivo
                            os.rename(caminho_completo, novo_caminho)
                            arquivos_extraidos.append(novo_caminho)
                            print(f"Arquivo extraído renomeado: {novo_nome}")
                        except Exception as e:
                            print(f"Erro ao renomear arquivo {arquivo} para {novo_nome}: {str(e)}")
                            arquivos_extraidos.append(caminho_completo)  # Adicionar o caminho original em caso de erro
        
        return arquivos_extraidos, pasta_temp
    
    except Exception as e:
        print(f"Erro ao extrair o arquivo ZIP {caminho_zip}: {str(e)}")
        # Se houve erro e criamos uma pasta temporária, tentar limpá-la
        if pasta_temp:
            try:
                shutil.rmtree(pasta_temp)
            except:
                pass
        return [], None

def processar_todos_arquivos():
    """
    Processa todos os arquivos TXT, CSV ou ZIP de transbordos nas pastas dados e dados/transbordos.
    Busca arquivos que começam com "RV Transbordo", "frente" e "transbordos" com extensão .csv, .txt ou .zip.
    Ignora arquivos que contenham "colhedora" no nome.
    """
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    
    # Diretórios para dados de entrada e saída
    diretorio_dados = os.path.join(diretorio_raiz, "dados")
    diretorio_transbordos = os.path.join(diretorio_raiz, "dados", "transbordos")
    diretorio_saida = os.path.join(diretorio_raiz, "output")
    
    # Verificar se os diretórios existem, caso contrário criar
    if not os.path.exists(diretorio_dados):
        os.makedirs(diretorio_dados)
    if not os.path.exists(diretorio_transbordos):
        os.makedirs(diretorio_transbordos)
    if not os.path.exists(diretorio_saida):
        os.makedirs(diretorio_saida)
    
    # Lista de diretórios para buscar arquivos
    diretorios_busca = [diretorio_dados, diretorio_transbordos]
    
    # Encontrar todos os arquivos TXT/CSV/ZIP de transbordos em ambos os diretórios
    arquivos = []
    arquivos_zip = []
    
    for diretorio in diretorios_busca:
        # Adicionar arquivos TXT sempre
        arquivos += glob.glob(os.path.join(diretorio, "RV Transbordo*.txt"))
        arquivos += glob.glob(os.path.join(diretorio, "*transbordo*.txt"))
        arquivos += glob.glob(os.path.join(diretorio, "frente*transbordos*.txt"))
        arquivos += glob.glob(os.path.join(diretorio, "transbordo*.txt"))
        
        # Adicionar arquivos CSV apenas se processCsv for True
        if processCsv:
            arquivos += glob.glob(os.path.join(diretorio, "RV Transbordo*.csv"))
            arquivos += glob.glob(os.path.join(diretorio, "*transbordo*.csv"))
            arquivos += glob.glob(os.path.join(diretorio, "frente*transbordos*.csv"))
            arquivos += glob.glob(os.path.join(diretorio, "transbordo*.csv"))
        
        # Adicionar arquivos ZIP
        arquivos_zip += glob.glob(os.path.join(diretorio, "RV Transbordo*.zip"))
        arquivos_zip += glob.glob(os.path.join(diretorio, "*transbordo*.zip"))
        arquivos_zip += glob.glob(os.path.join(diretorio, "frente*transbordos*.zip"))
        arquivos_zip += glob.glob(os.path.join(diretorio, "transbordo*.zip"))
    
    # Filtrar arquivos que contenham "colhedora" no nome (case insensitive)
    arquivos = [arquivo for arquivo in arquivos if "colhedora" not in os.path.basename(arquivo).lower()]
    arquivos_zip = [arquivo for arquivo in arquivos_zip if "colhedora" not in os.path.basename(arquivo).lower()]
    
    # Remover possíveis duplicatas
    arquivos = list(set(arquivos))
    arquivos_zip = list(set(arquivos_zip))
    
    if not arquivos and not arquivos_zip:
        print("Nenhum arquivo de transbordos encontrado nas pastas dados ou dados/transbordos!")
        return
    
    print(f"Encontrados {len(arquivos)} arquivos de transbordos (TXT/CSV) para processar.")
    print(f"Encontrados {len(arquivos_zip)} arquivos ZIP de transbordos para processar.")
    
    # Processar cada arquivo TXT/CSV
    for arquivo in arquivos:
        processar_arquivo(arquivo, diretorio_saida)
    
    # Processar cada arquivo ZIP
    for arquivo_zip in arquivos_zip:
        print(f"\nProcessando arquivo ZIP: {os.path.basename(arquivo_zip)}")
        
        # Extrair arquivo ZIP para pasta temporária
        arquivos_extraidos, pasta_temp = extrair_arquivo_zip(arquivo_zip)
        
        if not arquivos_extraidos:
            print(f"Nenhum arquivo TXT ou CSV encontrado no ZIP {os.path.basename(arquivo_zip)}")
            continue
        
        print(f"Extraídos {len(arquivos_extraidos)} arquivos do ZIP.")
        
        # Processar cada arquivo extraído
        for arquivo_extraido in arquivos_extraidos:
            # Filtrar arquivos que contenham "colhedora" no nome
            if "colhedora" not in os.path.basename(arquivo_extraido).lower():
                processar_arquivo(arquivo_extraido, diretorio_saida)
        
        # Limpar pasta temporária se foi criada
        if pasta_temp:
            try:
                shutil.rmtree(pasta_temp)
                print(f"Pasta temporária removida: {pasta_temp}")
            except Exception as e:
                print(f"Erro ao remover pasta temporária {pasta_temp}: {str(e)}")

def processar_arquivo(caminho_arquivo, diretorio_saida):
    """
    Processa um único arquivo e gera o Excel de saída.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo a ser processado
        diretorio_saida (str): Diretório onde o arquivo de saída será salvo
    """
    # Obter apenas o nome do arquivo (sem caminho e sem extensão)
    nome_base = os.path.splitext(os.path.basename(caminho_arquivo))[0]
    
    # Nome de saída igual ao original, mas com extensão .xlsx na pasta output
    arquivo_saida = os.path.join(diretorio_saida, f"{nome_base}.xlsx")
    
    print(f"\nProcessando arquivo: {os.path.basename(caminho_arquivo)}")
    print(f"Arquivo de saída: {os.path.basename(arquivo_saida)}")
    
    # Processar o arquivo base
    df_base = processar_arquivo_base(caminho_arquivo)
    if df_base is None:
        print(f"Erro ao processar {os.path.basename(caminho_arquivo)}. Pulando para o próximo arquivo.")
        return
    
    # Carregar substituições de operadores
    substituicoes = carregar_substituicoes_operadores()
    
    # Identificar operadores com IDs que começam com 133 e têm 7 dígitos
    mapeamento_duplicados, df_duplicados = identificar_operadores_duplicados(df_base, substituicoes)
    
    # Aplicar as substituições
    if substituicoes:
        df_base, df_substituicoes = aplicar_substituicao_operadores(df_base, substituicoes)
    else:
        df_substituicoes = pd.DataFrame(columns=['ID Original', 'Nome Original', 'ID Nova', 'Nome Novo', 'Registros Afetados'])
    
    # Se o DataFrame estiver vazio, gerar apenas a planilha BASE
    if len(df_base) == 0:
        writer = pd.ExcelWriter(arquivo_saida, engine='openpyxl')
        df_base.to_excel(writer, sheet_name='BASE', index=False)
        if not df_duplicados.empty:
            df_duplicados.to_excel(writer, sheet_name='IDs Encontradas', index=False)
        writer.close()
        print(f"Arquivo {arquivo_saida} gerado com apenas a planilha BASE (sem dados).")
        return
    
    # Calcular a Base Calculo
    base_calculo = calcular_base_calculo(df_base)
    
    # Calcular as métricas auxiliares
    disp_mecanica = calcular_disponibilidade_mecanica(df_base)
    eficiencia_energetica = calcular_eficiencia_energetica(base_calculo)
    motor_ocioso = calcular_motor_ocioso(base_calculo, df_base)
    falta_apontamento = calcular_falta_apontamento(base_calculo)
    uso_gps = calcular_uso_gps(base_calculo)
    horas_por_frota = calcular_horas_por_frota(df_base)
    
    # Calcular média de velocidade por operador
    media_velocidade = calcular_media_velocidade(df_base)
    
    # Criar o arquivo Excel com todas as planilhas
    criar_excel_com_planilhas(
        df_base, base_calculo, disp_mecanica, eficiencia_energetica,
        motor_ocioso, falta_apontamento, uso_gps, horas_por_frota, arquivo_saida,
        df_duplicados,  # Adicionar a tabela de IDs duplicadas
        media_velocidade,  # Adicionar a tabela de média de velocidade
        df_substituicoes  # Adicionar a tabela de IDs substituídas
    )
    
    print(f"Arquivo {arquivo_saida} gerado com sucesso!")

if __name__ == "__main__":
    print("="*80)
    print("Iniciando processamento de arquivos de transbordos...")
    print(f"Processamento de arquivos CSV: {'ATIVADO' if processCsv else 'DESATIVADO'}")
    print("Este script processa arquivos de transbordos e gera planilhas Excel com métricas")
    print("Suporta arquivos TXT, CSV e ZIP")
    print("Ignorando arquivos que contenham 'colhedora' no nome")
    print("="*80)
    processar_todos_arquivos()
    print("\nProcessamento concluído!") 
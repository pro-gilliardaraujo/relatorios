"""
Script para processamento completo de dados de monitoramento de colhedoras.
Lê arquivos TXT ou CSV na pasta raiz, processa-os e gera arquivos Excel com planilhas auxiliares prontas.
"""

import pandas as pd
import numpy as np
import os
import glob
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Configurações
processCsv = False  # Altere para True quando quiser processar arquivos CSV

# Constantes
COLUNAS_REMOVER = [
    'Justificativa Corte Base Desligado',
    'Latitude',
    'Longitude',
    'Regional',
    'Tipo de Equipamento',
    'Unidade',
    'Centro de Custo'
]

COLUNAS_DESEJADAS = [
    'Data', 'Hora', 'Equipamento', 'Apertura do Rolo', 'Codigo da Operacao',
    'Codigo Frente (digitada)', 'Corporativo', 'Corte Base Automatico/Manual',
    'Descricao Equipamento', 'Estado', 'Estado Operacional', 'Esteira Ligada',
    'Field Cruiser', 'Grupo Equipamento/Frente', 'Grupo Operacao', 'Horimetro',
    'Implemento Ligado', 'Motor Ligado', 'Operacao', 'Operador', 'Pressao de Corte',
    'RPM Extrator', 'RPM Motor', 'RTK (Piloto Automatico)', 'Fazenda', 'Zona',
    'Talhao', 'Velocidade', 'Diferença_Hora', 'Parada com Motor Ligado',
    'Horas Produtivas'
]

# Valores a serem filtrados
OPERADORES_EXCLUIR = ["9999 - TROCA DE TURNO"]

def processar_arquivo_base(caminho_arquivo):
    """
    Processa o arquivo TXT ou CSV e retorna o DataFrame com as transformações necessárias.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo TXT ou CSV de entrada
    
    Returns:
        DataFrame: DataFrame processado com todas as transformações
    """
    # Lista de codificações para tentar
    codificacoes = ['utf-8', 'latin1', 'ISO-8859-1', 'cp1252']
    
    for codificacao in codificacoes:
        try:
            # Leitura do arquivo (TXT ou CSV são tratados da mesma forma se usarem separador ';')
            df = pd.read_csv(caminho_arquivo, sep=';', encoding=codificacao)
            print(f"Arquivo lido com sucesso usando {codificacao}! Total de linhas: {len(df)}")
            
            # Verificar se o DataFrame está vazio (apenas cabeçalhos sem dados)
            if len(df) == 0:
                print(f"O arquivo {caminho_arquivo} contém apenas cabeçalhos sem dados.")
                # Retornar o DataFrame vazio mas com as colunas, em vez de None
                # Garantir que todas as colunas desejadas existam
                for col in COLUNAS_DESEJADAS:
                    if col not in df.columns:
                        df[col] = np.nan
                # Reorganizar as colunas na ordem desejada
                colunas_existentes = [col for col in COLUNAS_DESEJADAS if col in df.columns]
                colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS]
                return df[colunas_existentes + colunas_extras]
            
            # Limpeza de espaços extras nos nomes das colunas
            df.columns = df.columns.str.strip()
            
            # Verificar se 'Data/Hora' existe e processá-la
            if 'Data/Hora' in df.columns:
                df[['Data', 'Hora']] = df['Data/Hora'].str.split(' ', expand=True)
                df = df.drop(columns=['Data/Hora'])
            
            # Conversão e cálculo de diferenças de hora
            df['Hora'] = pd.to_datetime(df['Hora'], format='%H:%M:%S', errors='coerce')
            
            # Calcular a diferença de hora em segundos para maior precisão e depois converter para horas
            # Verificamos se a coluna 'Diferença_Hora' já existe
            if 'Diferença_Hora' not in df.columns or df['Diferença_Hora'].isna().any():
                df['Diferença_Hora'] = df['Hora'].diff().dt.total_seconds() / 3600
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if pd.isna(x) or x < 0 else x)
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if x > 0.50 else round(x, 4))
            else:
                # Certifica-se de que a coluna 'Diferença_Hora' esteja limpa de caracteres extra e como número
                df['Diferença_Hora'] = pd.to_numeric(df['Diferença_Hora'].astype(str).str.strip(), errors='coerce')
                df['Diferença_Hora'] = df['Diferença_Hora'].fillna(0)
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if x > 0.50 else round(x, 4))
            
            # Cálculos adicionais
            RPM_MINIMO = 300  # Definindo constante para RPM mínimo
            if 'Parada com Motor Ligado' not in df.columns:
                df['Parada com Motor Ligado'] = ((df['Velocidade'] == 0) & 
                                               (df['RPM Motor'] >= RPM_MINIMO)).astype(int)
            
            # Verifica se Horas Produtivas já existe
            if 'Horas Produtivas' not in df.columns or df['Horas Produtivas'].isna().any():
                # Calcular horas produtivas com alta precisão
                df['Horas Produtivas'] = df.apply(
                    lambda row: round(row['Diferença_Hora'], 4) if row['Grupo Operacao'] == 'Produtiva' else 0,
                    axis=1
                )
            else:
                # Limpa e converte para número
                df['Horas Produtivas'] = pd.to_numeric(df['Horas Produtivas'].astype(str).str.strip(), errors='coerce')
                df['Horas Produtivas'] = df['Horas Produtivas'].fillna(0)
            
            # Conversão de colunas binárias para valores numéricos (garantindo que sejam números)
            for col in ['Esteira Ligada', 'Motor Ligado', 'Field Cruiser', 'RTK (Piloto Automatico)', 'Implemento Ligado']:
                if col in df.columns:
                    # Se a coluna for texto (LIGADO/DESLIGADO), converter para 1/0
                    if df[col].dtype == 'object':
                        df[col] = df[col].replace({'LIGADO': 1, 'DESLIGADO': 0})
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
            
            # Limpeza e organização das colunas
            df = df.drop(columns=COLUNAS_REMOVER, errors='ignore')
            
            # Garantir que todas as colunas desejadas existam
            for col in COLUNAS_DESEJADAS:
                if col not in df.columns:
                    df[col] = np.nan
            
            # Reorganizar as colunas na ordem desejada
            colunas_existentes = [col for col in COLUNAS_DESEJADAS if col in df.columns]
            colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS]
            df = df[colunas_existentes + colunas_extras]
            
            return df
            
        except UnicodeDecodeError:
            print(f"Tentativa com codificação {codificacao} falhou, tentando próxima codificação...")
            continue
        except Exception as e:
            print(f"Erro ao processar o arquivo com codificação {codificacao}: {str(e)}")
            continue
    
    # Se chegou aqui, todas as tentativas de codificação falharam
    print(f"Erro: Não foi possível ler o arquivo {caminho_arquivo} com nenhuma das codificações tentadas.")
    return None

def calcular_base_calculo(df):
    """
    Calcula a tabela de Base Calculo a partir do DataFrame processado.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Tabela Base Calculo com todas as métricas calculadas
    """
    # Extrair combinações únicas de Equipamento, Grupo Equipamento/Frente e Operador
    combinacoes = df[['Equipamento', 'Grupo Equipamento/Frente', 'Operador']].drop_duplicates().reset_index(drop=True)
    
    # Filtrar operadores excluídos
    combinacoes = combinacoes[~combinacoes['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Inicializar as colunas de métricas
    resultados = []
    
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    # Calcular as métricas para cada combinação
    for idx, row in combinacoes.iterrows():
        equipamento = row['Equipamento']
        grupo = row['Grupo Equipamento/Frente']
        operador = row['Operador']
        
        # Filtrar dados para esta combinação
        filtro = (df['Equipamento'] == equipamento) & \
                (df['Grupo Equipamento/Frente'] == grupo) & \
                (df['Operador'] == operador)
        
        dados_filtrados = df[filtro]
        
        # Horas totais - manter mais casas decimais para cálculos intermediários
        horas_totais = round(dados_filtrados['Diferença_Hora'].sum(), 4)
        
        # Horas elevador (Esteira Ligada = 1 E Pressão de Corte > 400)
        horas_elevador = round(dados_filtrados[
            (dados_filtrados['Esteira Ligada'] == 1) & 
            (dados_filtrados['Pressao de Corte'] > 400)
        ]['Diferença_Hora'].sum(), 4)
        
        # Percentual horas elevador (em decimal 0-1)
        percent_elevador = calcular_porcentagem(horas_elevador, horas_totais)
        
        # RTK (Piloto Automático = 1 e Field Cruiser = 1)
        rtk = round(dados_filtrados[(dados_filtrados['RTK (Piloto Automatico)'] == 1) & 
                             (dados_filtrados['Field Cruiser'] == 1)]['Diferença_Hora'].sum(), 4)
        
        # Horas Produtivas
        horas_produtivas = round(dados_filtrados['Horas Produtivas'].sum(), 4)
        
        # % Utilização RTK (em decimal 0-1)
        utilizacao_rtk = calcular_porcentagem(rtk, horas_produtivas)
        
        # Motor Ligado
        motor_ligado = round(dados_filtrados[dados_filtrados['Motor Ligado'] == 1]['Diferença_Hora'].sum(), 4)
        
        # % Eficiência Elevador (em decimal 0-1)
        eficiencia_elevador = calcular_porcentagem(horas_elevador, motor_ligado)
        
        # Parado com Motor Ligado
        parado_motor_ligado = round(dados_filtrados[dados_filtrados['Parada com Motor Ligado'] == 1]['Diferença_Hora'].sum(), 4)
        
        # % Parado com motor ligado (em decimal 0-1)
        percent_parado_motor = calcular_porcentagem(parado_motor_ligado, motor_ligado)
        
        resultados.append({
            'Equipamento': equipamento,
            'Grupo Equipamento/Frente': grupo,
            'Operador': operador,
            'Horas totais': horas_totais,
            'Horas elevador': horas_elevador,
            '%': percent_elevador,
            'RTK': rtk,
            'Horas Produtivas': horas_produtivas,
            '% Utilização RTK': utilizacao_rtk,
            'Motor Ligado': motor_ligado,
            '% Eficiência Elevador': eficiencia_elevador,
            'Parado Com Motor Ligado': parado_motor_ligado,
            '% Parado com motor ligado': percent_parado_motor
        })
    
    return pd.DataFrame(resultados)

def calcular_disponibilidade_mecanica(df):
    """
    Calcula a disponibilidade mecânica para cada equipamento.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Disponibilidade mecânica por equipamento
    """
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    # Filtramos os dados excluindo os operadores da lista
    df_filtrado = df[~df['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Agrupar por Equipamento e calcular horas por grupo operacional
    equipamentos = df_filtrado['Equipamento'].unique()
    resultados = []
    
    for equipamento in equipamentos:
        dados_equip = df_filtrado[df_filtrado['Equipamento'] == equipamento]
        total_horas = round(dados_equip['Diferença_Hora'].sum(), 4)
        
        # Calcular horas de manutenção
        manutencao = round(dados_equip[dados_equip['Grupo Operacao'] == 'Manutenção']['Diferença_Hora'].sum(), 4)
        
        # A disponibilidade mecânica é o percentual de tempo fora de manutenção
        disp_mecanica = calcular_porcentagem(total_horas - manutencao, total_horas)
        
        resultados.append({
            'Frota': equipamento,
            'Disponibilidade': disp_mecanica
        })
    
    return pd.DataFrame(resultados)

def calcular_horas_por_frota(df):
    """
    Calcula o total de horas registradas para cada frota e a diferença para 24 horas.
    Esta função NÃO aplica qualquer filtro de operador.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Horas totais por frota
    """
    # Agrupar por Equipamento e somar as diferenças de hora
    equipamentos = df['Equipamento'].unique()
    resultados = []
    
    for equipamento in equipamentos:
        dados_equip = df[df['Equipamento'] == equipamento]
        total_horas = round(dados_equip['Diferença_Hora'].sum(), 2)
        
        # Calcular a diferença para 24 horas
        diferenca_24h = round(max(24 - total_horas, 0), 2)
        
        resultados.append({
            'Frota': equipamento,
            'Horas Registradas': total_horas,
            'Diferença para 24h': diferenca_24h
        })
    
    return pd.DataFrame(resultados)

def calcular_eficiencia_energetica(base_calculo):
    """
    Calcula a eficiência energética por operador.
    Eficiência energética = Horas elevador / Horas motor ligado
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Eficiência energética por operador
    """
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    # Agrupar por operador (já filtrado pela função calcular_base_calculo)
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Eficiência Energética = horas elevador / motor ligado
        horas_elevador_sum = round(dados_op['Horas elevador'].sum(), 4)
        motor_ligado_sum = round(dados_op['Motor Ligado'].sum(), 4)
        
        # Calcular eficiência - já está em decimal, não precisa multiplicar por 100
        eficiencia = calcular_porcentagem(horas_elevador_sum, motor_ligado_sum)
        
        # Garantir que não ultrapasse 100%
        eficiencia = min(eficiencia, 1.0)
        
        resultados.append({
            'Operador': operador,
            'Eficiência': eficiencia
        })
    
    return pd.DataFrame(resultados)

def calcular_hora_elevador(df, base_calculo):
    """
    Calcula as horas de elevador por operador.
    Considera-se horas de elevador quando:
    - Esteira Ligada = 1
    - Pressão de Corte > 400
    
    Args:
        df (DataFrame): DataFrame base processado
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Horas de elevador por operador
    """
    # Agrupar por operador (já filtrado pela função calcular_base_calculo)
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Somar horas de elevador da base de cálculo (já filtradas corretamente)
        horas_elevador_sum = round(dados_op['Horas elevador'].sum(), 2)
        
        resultados.append({
            'Operador': operador,
            'Horas': horas_elevador_sum
        })
    
    return pd.DataFrame(resultados)

def calcular_motor_ocioso(base_calculo):
    """
    Calcula o percentual de motor ocioso por operador.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de motor ocioso por operador
    """
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    # Agrupar por operador (já filtrado pela função calcular_base_calculo)
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Motor Ocioso = Parado Com Motor Ligado / Motor Ligado
        parado_motor_sum = round(dados_op['Parado Com Motor Ligado'].sum(), 4)
        motor_ligado_sum = round(dados_op['Motor Ligado'].sum(), 4)
        
        percentual = calcular_porcentagem(parado_motor_sum, motor_ligado_sum)
        
        resultados.append({
            'Operador': operador,
            'Porcentagem': percentual
        })
    
    return pd.DataFrame(resultados)

def calcular_uso_gps(df, base_calculo):
    """
    Calcula o percentual de uso de GPS por operador.
    O uso de GPS é definido como o tempo em que o equipamento está:
    - Estado é "TRABALHANDO" ou "COLHEITA" 
    - Piloto Automático (RTK) = 1
    - Velocidade > 0
    
    Args:
        df (DataFrame): DataFrame base processado
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de uso de GPS por operador
    """
    # Função para calcular valores com alta precisão e depois formatar
    def calcular_porcentagem(numerador, denominador, precisao=4):
        """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
        if denominador > 0:
            return round((numerador / denominador), precisao)
        return 0.0
    
    # Agrupar por operador (já filtrado pela função calcular_base_calculo)
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados base para este operador e grupo
        filtro_base = (df['Operador'] == operador) & \
                      (df['Grupo Equipamento/Frente'] == grupo)
        dados_op_base = df[filtro_base]
        
        # Calcular tempo total trabalhando
        tempo_trabalhando = round(dados_op_base[
            (dados_op_base['Estado'].isin(['TRABALHANDO', 'COLHEITA']))
        ]['Diferença_Hora'].sum(), 4)
        
        # Calcular tempo com GPS ativo (condições combinadas)
        tempo_gps_ativo = round(dados_op_base[
            (dados_op_base['Estado'].isin(['TRABALHANDO', 'COLHEITA'])) &
            (dados_op_base['RTK (Piloto Automatico)'] == 1) &
            (dados_op_base['Velocidade'] > 0)
        ]['Diferença_Hora'].sum(), 4)
        
        # Calcular percentual em formato decimal (0-1)
        percentual = calcular_porcentagem(tempo_gps_ativo, tempo_trabalhando)
        
        # Garantir que não ultrapasse 100% (1.0)
        percentual = min(percentual, 1.0)
        
        resultados.append({
            'Operador': operador,
            'Porcentagem': percentual
        })
    
    return pd.DataFrame(resultados)

def criar_excel_com_planilhas(df_base, base_calculo, disp_mecanica, eficiencia_energetica, 
                             hora_elevador, motor_ocioso, uso_gps, horas_por_frota, caminho_saida):
    """
    Cria um arquivo Excel com todas as planilhas auxiliares.
    
    Args:
        df_base (DataFrame): DataFrame base processado
        base_calculo (DataFrame): Tabela Base Calculo
        disp_mecanica (DataFrame): Disponibilidade mecânica
        eficiencia_energetica (DataFrame): Eficiência energética
        hora_elevador (DataFrame): Horas de elevador
        motor_ocioso (DataFrame): Motor ocioso
        uso_gps (DataFrame): Uso GPS
        horas_por_frota (DataFrame): Horas totais registradas por frota
        caminho_saida (str): Caminho do arquivo Excel de saída
    """
    writer = pd.ExcelWriter(caminho_saida, engine='openpyxl')
    
    # Arredondamento fixo para 2 casas decimais em todas as colunas numéricas antes de exportar
    # Base Calculo - garantir que todas as colunas numéricas tenham 2 casas decimais
    colunas_numericas = ['Horas totais', 'Horas elevador', '%', 'RTK', 'Horas Produtivas', 
                         '% Utilização RTK', 'Motor Ligado', '% Eficiência Elevador', 
                         'Parado Com Motor Ligado', '% Parado com motor ligado']
    
    for col in colunas_numericas:
        if col in base_calculo.columns:
            base_calculo[col] = base_calculo[col].apply(lambda x: round(x, 2))
    
    # Arredondar valores nas outras planilhas
    disp_mecanica['Disponibilidade'] = disp_mecanica['Disponibilidade'].apply(lambda x: round(x, 4))
    eficiencia_energetica['Eficiência'] = eficiencia_energetica['Eficiência'].apply(lambda x: round(x, 4))
    hora_elevador['Horas'] = hora_elevador['Horas'].apply(lambda x: round(x, 2))
    motor_ocioso['Porcentagem'] = motor_ocioso['Porcentagem'].apply(lambda x: round(x, 4))
    uso_gps['Porcentagem'] = uso_gps['Porcentagem'].apply(lambda x: round(x, 4))
    horas_por_frota['Horas Registradas'] = horas_por_frota['Horas Registradas'].apply(lambda x: round(x, 2))
    horas_por_frota['Diferença para 24h'] = horas_por_frota['Diferença para 24h'].apply(lambda x: round(x, 2))
    
    # Salvar cada DataFrame em uma planilha separada
    df_base.to_excel(writer, sheet_name='BASE', index=False)
    base_calculo.to_excel(writer, sheet_name='Base Calculo', index=False)
    
    # Planilhas auxiliares (formatadas conforme necessário)
    disp_mecanica.to_excel(writer, sheet_name='1_Disponibilidade Mecânica', index=False)
    eficiencia_energetica.to_excel(writer, sheet_name='2_Eficiência Energética', index=False)
    hora_elevador.to_excel(writer, sheet_name='3_Hora Elevador', index=False)
    motor_ocioso.to_excel(writer, sheet_name='4_Motor Ocioso', index=False)
    uso_gps.to_excel(writer, sheet_name='5_Uso GPS', index=False)
    horas_por_frota.to_excel(writer, sheet_name='Horas por Frota', index=False)
    
    # Aplicar formatação nas planilhas
    workbook = writer.book
    
    # Formatar planilha de Disponibilidade Mecânica
    worksheet = workbook['1_Disponibilidade Mecânica']
    for row in range(2, worksheet.max_row + 1):  # Começando da linha 2 (ignorando cabeçalho)
        cell = worksheet.cell(row=row, column=2)  # Coluna B (Disponibilidade)
        cell.number_format = '0.00%'  # Formato de porcentagem com 2 casas
    
    # Formatar planilha de Eficiência Energética
    worksheet = workbook['2_Eficiência Energética']
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=2)  # Coluna B (Eficiência)
        cell.number_format = '0.00%'  # Formato de porcentagem com 2 casas
    
    # Formatar planilha de Hora Elevador
    worksheet = workbook['3_Hora Elevador']
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=2)  # Coluna B (Horas)
        cell.number_format = '0.00'  # Formato decimal com 2 casas
    
    # Formatar planilha de Motor Ocioso
    worksheet = workbook['4_Motor Ocioso']
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
        cell.number_format = '0.00%'  # Formato de porcentagem com 2 casas
    
    # Formatar planilha de Uso GPS
    worksheet = workbook['5_Uso GPS']
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
        cell.number_format = '0.00%'  # Formato de porcentagem com 2 casas
    
    # Formatar planilha de Base Calculo
    worksheet = workbook['Base Calculo']
    for row in range(2, worksheet.max_row + 1):
        # Formatar colunas decimais
        columns_decimal = [4, 5, 7, 8, 10, 12]  # Colunas D, E, G, H, J, L (Horas totais, Horas elevador, etc.)
        for col in columns_decimal:
            if col <= worksheet.max_column:
                cell = worksheet.cell(row=row, column=col)
                cell.number_format = '0.00'  # Formato decimal com 2 casas
        
        # Formatar colunas de porcentagem
        columns_percent = [6, 9, 11, 13]  # Colunas F, I, K, M (%, % Utilização RTK, etc.)
        for col in columns_percent:
            if col <= worksheet.max_column:
                cell = worksheet.cell(row=row, column=col)
                cell.number_format = '0.00%'  # Formato de porcentagem com 2 casas
    
    # Formatar planilha de Horas por Frota
    worksheet = workbook['Horas por Frota']
    for row in range(2, worksheet.max_row + 1):
        # Coluna B (Horas Registradas)
        cell_b = worksheet.cell(row=row, column=2)
        cell_b.number_format = '0.00'  # Formato decimal com 2 casas
        
        # Coluna C (Diferença para 24h)
        cell_c = worksheet.cell(row=row, column=3)
        cell_c.number_format = '0.00'  # Formato decimal com 2 casas
    
    writer.close()
    print(f"Arquivo Excel salvo com sucesso em {caminho_saida}")

def processar_todos_arquivos():
    """
    Processa todos os arquivos TXT ou CSV de colhedoras nas pastas dados e dados/colhedoras.
    Ignora arquivos que contenham "transbordo" no nome.
    """
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    
    # Diretórios para dados de entrada e saída
    diretorio_dados = os.path.join(diretorio_raiz, "dados")
    diretorio_colhedoras = os.path.join(diretorio_raiz, "dados", "colhedoras")
    diretorio_saida = os.path.join(diretorio_raiz, "output")
    
    # Verificar se os diretórios existem, caso contrário criar
    if not os.path.exists(diretorio_dados):
        os.makedirs(diretorio_dados)
    if not os.path.exists(diretorio_colhedoras):
        os.makedirs(diretorio_colhedoras)
    if not os.path.exists(diretorio_saida):
        os.makedirs(diretorio_saida)
    
    # Lista de diretórios para buscar arquivos
    diretorios_busca = [diretorio_dados, diretorio_colhedoras]
    
    # Encontrar todos os arquivos de colhedoras em ambos os diretórios
    arquivos = []
    
    for diretorio in diretorios_busca:
        # Adicionar arquivos TXT sempre
        arquivos += glob.glob(os.path.join(diretorio, "RV Colhedora*.txt"))
        arquivos += glob.glob(os.path.join(diretorio, "*colhedora*.txt"))
        arquivos += glob.glob(os.path.join(diretorio, "colhedora*.txt"))
        
        # Adicionar arquivos CSV apenas se processCsv for True
        if processCsv:
            arquivos += glob.glob(os.path.join(diretorio, "RV Colhedora*.csv"))
            arquivos += glob.glob(os.path.join(diretorio, "*colhedora*.csv"))
            arquivos += glob.glob(os.path.join(diretorio, "colhedora*.csv"))
    
    # Filtrar arquivos que contenham "transbordo" no nome (case insensitive)
    arquivos = [arquivo for arquivo in arquivos if "transbordo" not in os.path.basename(arquivo).lower()]
    
    # Remover possíveis duplicatas
    arquivos = list(set(arquivos))
    
    if not arquivos:
        print("Nenhum arquivo de colhedoras encontrado nas pastas dados ou dados/colhedoras!")
        return
    
    print(f"Encontrados {len(arquivos)} arquivos de colhedoras para processar.")
    
    # Processar cada arquivo
    for arquivo in arquivos:
        # Obter apenas o nome do arquivo (sem caminho e sem extensão)
        nome_base = os.path.splitext(os.path.basename(arquivo))[0]
        
        # Nome de saída igual ao original, mas com extensão .xlsx na pasta output
        arquivo_saida = os.path.join(diretorio_saida, f"{nome_base}.xlsx")
        
        print(f"\nProcessando arquivo: {os.path.basename(arquivo)}")
        print(f"Arquivo de saída: {os.path.basename(arquivo_saida)}")
        
        # Processar o arquivo base
        df_base = processar_arquivo_base(arquivo)
        if df_base is None:
            print(f"Erro ao processar {os.path.basename(arquivo)}. Pulando para o próximo arquivo.")
            continue
        
        # Se o DataFrame estiver vazio, gerar apenas a planilha BASE
        if len(df_base) == 0:
            writer = pd.ExcelWriter(arquivo_saida, engine='openpyxl')
            df_base.to_excel(writer, sheet_name='BASE', index=False)
            writer.close()
            print(f"Arquivo {arquivo_saida} gerado com apenas a planilha BASE (sem dados).")
            continue
        
        # Calcular a Base Calculo
        base_calculo = calcular_base_calculo(df_base)
        
        # Calcular as métricas auxiliares
        disp_mecanica = calcular_disponibilidade_mecanica(df_base)
        eficiencia_energetica = calcular_eficiencia_energetica(base_calculo)
        hora_elevador = calcular_hora_elevador(df_base, base_calculo)
        motor_ocioso = calcular_motor_ocioso(base_calculo)
        uso_gps = calcular_uso_gps(df_base, base_calculo)
        horas_por_frota = calcular_horas_por_frota(df_base)
        
        # Criar o arquivo Excel com todas as planilhas
        criar_excel_com_planilhas(
            df_base, base_calculo, disp_mecanica, eficiencia_energetica,
            hora_elevador, motor_ocioso, uso_gps, horas_por_frota, arquivo_saida
        )
        
        print(f"Arquivo {arquivo_saida} gerado com sucesso!")

if __name__ == "__main__":
    print("="*80)
    print("Iniciando processamento de arquivos de colhedoras...")
    print(f"Processamento de arquivos CSV: {'ATIVADO' if processCsv else 'DESATIVADO'}")
    print("Este script processa arquivos de colhedoras e gera planilhas Excel com métricas")
    print("Ignorando arquivos que contenham 'transbordo' no nome")
    print("="*80)
    processar_todos_arquivos()
    print("\nProcessamento concluído!") 
"""
Script para processamento unificado de dados de monitoramento de colhedoras e transbordos.
Lê arquivos TXT ou CSV nas pastas especificadas, processa-os e gera um único arquivo Excel 
com todas as planilhas, adicionando prefixos "CD_" para colhedoras e "TT_" para transbordos.
"""

import pandas as pd
import numpy as np
import os
import glob
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Configurações
processCsv = True  # Altere para True quando quiser processar arquivos CSV
arquivo_saida_unificado = "relatorio_unificado.xlsx"  # Nome do arquivo de saída unificado

# Constantes comuns
OPERADORES_EXCLUIR = ["9999 - TROCA DE TURNO"]

# Constantes para colhedoras
COLUNAS_REMOVER_COLHEDORAS = [
    'Justificativa Corte Base Desligado',
    'Latitude',
    'Longitude',
    'Regional',
    'Tipo de Equipamento',
    'Unidade',
    'Centro de Custo'
]

COLUNAS_DESEJADAS_COLHEDORAS = [
    'Data', 'Hora', 'Equipamento', 'Apertura do Rolo', 'Codigo da Operacao',
    'Codigo Frente (digitada)', 'Corporativo', 'Corte Base Automatico/Manual',
    'Descricao Equipamento', 'Estado', 'Estado Operacional', 'Esteira Ligada',
    'Field Cruiser', 'Grupo Equipamento/Frente', 'Grupo Operacao', 'Horimetro',
    'Implemento Ligado', 'Motor Ligado', 'Operacao', 'Operador', 'Pressao de Corte',
    'RPM Extrator', 'RPM Motor', 'RTK (Piloto Automatico)', 'Fazenda', 'Zona',
    'Talhao', 'Velocidade', 'Diferença_Hora', 'Parada com Motor Ligado',
    'Horas Produtivas'
]

# Constantes para transbordos
COLUNAS_REMOVER_TRANSBORDOS = [
    'Latitude',
    'Longitude',
    'Regional',
    'Unidade',
    'Centro de Custo'
]

COLUNAS_DESEJADAS_TRANSBORDOS = [
    'Data', 'Hora', 'Equipamento', 'Codigo da Operacao',
    'Codigo Frente (digitada)', 'Corporativo',
    'Descricao Equipamento', 'Estado', 'Estado Operacional',
    'Grupo Equipamento/Frente', 'Grupo Operacao', 'Horimetro',
    'Motor Ligado', 'Operacao', 'Operador',
    'RPM Motor', 'Fazenda', 'Zona',
    'Talhao', 'Velocidade', 'Diferença_Hora', 'Parado Com Motor Ligado',
    'Horas Produtivas', 'GPS'
]

# Funções comuns
def calcular_porcentagem(numerador, denominador, precisao=4):
    """Calcula porcentagem como decimal (0-1) evitando divisão por zero."""
    if denominador > 0:
        return round((numerador / denominador), precisao)
    return 0.0

def calcular_disponibilidade_mecanica(df):
    """
    Calcula a disponibilidade mecânica para cada equipamento.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Disponibilidade mecânica por equipamento
    """
    # Filtramos os dados excluindo os operadores da lista
    df_filtrado = df[~df['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Agrupar por Equipamento e calcular horas por grupo operacional
    equipamentos = df_filtrado['Equipamento'].unique()
    resultados = []
    
    for equipamento in equipamentos:
        dados_equip = df_filtrado[df_filtrado['Equipamento'] == equipamento]
        total_horas = round(dados_equip['Diferença_Hora'].sum(), 4)
        
        # Calcular horas de manutenção
        manutencao = round(dados_equip[dados_equip['Grupo Operacao'] == 'Manutenção']['Diferença_Hora'].sum(), 4)
        
        # A disponibilidade mecânica é o percentual de tempo fora de manutenção
        disp_mecanica = calcular_porcentagem(total_horas - manutencao, total_horas)
        
        resultados.append({
            'Frota': equipamento,
            'Disponibilidade': disp_mecanica
        })
    
    return pd.DataFrame(resultados)

def calcular_horas_por_frota(df):
    """
    Calcula o total de horas registradas para cada frota e a diferença para 24 horas.
    Esta função NÃO aplica qualquer filtro de operador.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Horas totais por frota
    """
    # Agrupar por Equipamento e somar as diferenças de hora
    equipamentos = df['Equipamento'].unique()
    resultados = []
    
    for equipamento in equipamentos:
        dados_equip = df[df['Equipamento'] == equipamento]
        total_horas = round(dados_equip['Diferença_Hora'].sum(), 2)
        
        # Calcular a diferença para 24 horas
        diferenca_24h = round(max(24 - total_horas, 0), 2)
        
        resultados.append({
            'Frota': equipamento,
            'Horas Registradas': total_horas,
            'Diferença para 24h': diferenca_24h
        })
    
    return pd.DataFrame(resultados)

# Funções para colhedoras
def processar_arquivo_colhedora(caminho_arquivo):
    """
    Processa o arquivo TXT ou CSV e retorna o DataFrame com as transformações necessárias para colhedoras.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo TXT ou CSV de entrada
    
    Returns:
        DataFrame: DataFrame processado com todas as transformações
    """
    # Lista de codificações para tentar
    codificacoes = ['utf-8', 'latin1', 'ISO-8859-1', 'cp1252']
    
    for codificacao in codificacoes:
        try:
            # Leitura do arquivo
            df = pd.read_csv(caminho_arquivo, sep=';', encoding=codificacao)
            print(f"Arquivo lido com sucesso usando {codificacao}! Total de linhas: {len(df)}")
            
            # Verificar se o DataFrame está vazio
            if len(df) == 0:
                print(f"O arquivo {caminho_arquivo} contém apenas cabeçalhos sem dados.")
                for col in COLUNAS_DESEJADAS_COLHEDORAS:
                    if col not in df.columns:
                        df[col] = np.nan
                colunas_existentes = [col for col in COLUNAS_DESEJADAS_COLHEDORAS if col in df.columns]
                colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS_COLHEDORAS]
                return df[colunas_existentes + colunas_extras]
            
            # Limpeza de espaços extras nos nomes das colunas
            df.columns = df.columns.str.strip()
            
            # Verificar se 'Data/Hora' existe e processá-la
            if 'Data/Hora' in df.columns:
                df[['Data', 'Hora']] = df['Data/Hora'].str.split(' ', expand=True)
                df = df.drop(columns=['Data/Hora'])
            
            # Conversão e cálculo de diferenças de hora
            df['Hora'] = pd.to_datetime(df['Hora'], format='%H:%M:%S', errors='coerce')
            
            # Calcular a diferença de hora em segundos para maior precisão e depois converter para horas
            if 'Diferença_Hora' not in df.columns or df['Diferença_Hora'].isna().any():
                df['Diferença_Hora'] = df['Hora'].diff().dt.total_seconds() / 3600
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if pd.isna(x) or x < 0 else x)
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if x > 0.50 else round(x, 4))
            else:
                df['Diferença_Hora'] = pd.to_numeric(df['Diferença_Hora'].astype(str).str.strip(), errors='coerce')
                df['Diferença_Hora'] = df['Diferença_Hora'].fillna(0)
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if x > 0.50 else round(x, 4))
            
            # Cálculos adicionais
            RPM_MINIMO = 300
            if 'Parada com Motor Ligado' not in df.columns:
                df['Parada com Motor Ligado'] = ((df['Velocidade'] == 0) & 
                                              (df['RPM Motor'] >= RPM_MINIMO)).astype(int)
            
            # Verificar se Horas Produtivas já existe
            if 'Horas Produtivas' not in df.columns or df['Horas Produtivas'].isna().any():
                df['Horas Produtivas'] = df.apply(
                    lambda row: round(row['Diferença_Hora'], 4) if row['Grupo Operacao'] == 'Produtiva' else 0,
                    axis=1
                )
            else:
                df['Horas Produtivas'] = pd.to_numeric(df['Horas Produtivas'].astype(str).str.strip(), errors='coerce')
                df['Horas Produtivas'] = df['Horas Produtivas'].fillna(0)
            
            # Conversão de colunas binárias para valores numéricos
            for col in ['Esteira Ligada', 'Motor Ligado', 'Field Cruiser', 'RTK (Piloto Automatico)', 'Implemento Ligado']:
                if col in df.columns:
                    if df[col].dtype == 'object':
                        df[col] = df[col].replace({'LIGADO': 1, 'DESLIGADO': 0})
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
            
            # Limpeza e organização das colunas
            df = df.drop(columns=COLUNAS_REMOVER_COLHEDORAS, errors='ignore')
            
            # Garantir que todas as colunas desejadas existam
            for col in COLUNAS_DESEJADAS_COLHEDORAS:
                if col not in df.columns:
                    df[col] = np.nan
            
            # Reorganizar as colunas na ordem desejada
            colunas_existentes = [col for col in COLUNAS_DESEJADAS_COLHEDORAS if col in df.columns]
            colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS_COLHEDORAS]
            df = df[colunas_existentes + colunas_extras]
            
            return df
            
        except UnicodeDecodeError:
            print(f"Tentativa com codificação {codificacao} falhou, tentando próxima codificação...")
            continue
        except Exception as e:
            print(f"Erro ao processar o arquivo com codificação {codificacao}: {str(e)}")
            continue
    
    print(f"Erro: Não foi possível ler o arquivo {caminho_arquivo} com nenhuma das codificações tentadas.")
    return None

def calcular_base_calculo_colhedora(df):
    """
    Calcula a tabela de Base Calculo a partir do DataFrame processado para colhedoras.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Tabela Base Calculo com todas as métricas calculadas
    """
    # Extrair combinações únicas de Equipamento, Grupo Equipamento/Frente e Operador
    combinacoes = df[['Equipamento', 'Grupo Equipamento/Frente', 'Operador']].drop_duplicates().reset_index(drop=True)
    
    # Filtrar operadores excluídos
    combinacoes = combinacoes[~combinacoes['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Inicializar as colunas de métricas
    resultados = []
    
    # Calcular as métricas para cada combinação
    for idx, row in combinacoes.iterrows():
        equipamento = row['Equipamento']
        grupo = row['Grupo Equipamento/Frente']
        operador = row['Operador']
        
        # Filtrar dados para esta combinação
        filtro = (df['Equipamento'] == equipamento) & \
                (df['Grupo Equipamento/Frente'] == grupo) & \
                (df['Operador'] == operador)
        
        dados_filtrados = df[filtro]
        
        # Horas totais - manter mais casas decimais para cálculos intermediários
        horas_totais = round(dados_filtrados['Diferença_Hora'].sum(), 4)
        
        # Horas elevador (Esteira Ligada = 1 E Pressão de Corte > 400)
        horas_elevador = round(dados_filtrados[
            (dados_filtrados['Esteira Ligada'] == 1) & 
            (dados_filtrados['Pressao de Corte'] > 400)
        ]['Diferença_Hora'].sum(), 4)
        
        # Percentual horas elevador (em decimal 0-1)
        percent_elevador = calcular_porcentagem(horas_elevador, horas_totais)
        
        # RTK (Piloto Automático = 1 e Field Cruiser = 1)
        rtk = round(dados_filtrados[(dados_filtrados['RTK (Piloto Automatico)'] == 1) & 
                             (dados_filtrados['Field Cruiser'] == 1)]['Diferença_Hora'].sum(), 4)
        
        # Horas Produtivas
        horas_produtivas = round(dados_filtrados['Horas Produtivas'].sum(), 4)
        
        # % Utilização RTK (em decimal 0-1)
        utilizacao_rtk = calcular_porcentagem(rtk, horas_produtivas)
        
        # Motor Ligado
        motor_ligado = round(dados_filtrados[dados_filtrados['Motor Ligado'] == 1]['Diferença_Hora'].sum(), 4)
        
        # % Eficiência Elevador (em decimal 0-1)
        eficiencia_elevador = calcular_porcentagem(horas_elevador, motor_ligado)
        
        # Parado com Motor Ligado
        parado_motor_ligado = round(dados_filtrados[dados_filtrados['Parada com Motor Ligado'] == 1]['Diferença_Hora'].sum(), 4)
        
        # % Parado com motor ligado (em decimal 0-1)
        percent_parado_motor = calcular_porcentagem(parado_motor_ligado, motor_ligado)
        
        resultados.append({
            'Equipamento': equipamento,
            'Grupo Equipamento/Frente': grupo,
            'Operador': operador,
            'Horas totais': horas_totais,
            'Horas elevador': horas_elevador,
            '%': percent_elevador,
            'RTK': rtk,
            'Horas Produtivas': horas_produtivas,
            '% Utilização RTK': utilizacao_rtk,
            'Motor Ligado': motor_ligado,
            '% Eficiência Elevador': eficiencia_elevador,
            'Parado Com Motor Ligado': parado_motor_ligado,
            '% Parado com motor ligado': percent_parado_motor
        })
    
    return pd.DataFrame(resultados)

def calcular_eficiencia_energetica_colhedora(base_calculo):
    """
    Calcula a eficiência energética por operador para colhedoras.
    Eficiência energética = Horas elevador / Horas motor ligado
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Eficiência energética por operador
    """
    # Agrupar por operador
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Eficiência Energética = horas elevador / motor ligado
        horas_elevador_sum = round(dados_op['Horas elevador'].sum(), 4)
        motor_ligado_sum = round(dados_op['Motor Ligado'].sum(), 4)
        
        # Calcular eficiência - já está em decimal, não precisa multiplicar por 100
        eficiencia = calcular_porcentagem(horas_elevador_sum, motor_ligado_sum)
        
        # Garantir que não ultrapasse 100%
        eficiencia = min(eficiencia, 1.0)
        
        resultados.append({
            'Operador': operador,
            'Eficiência': eficiencia
        })
    
    return pd.DataFrame(resultados)

def calcular_hora_elevador(df, base_calculo):
    """
    Calcula as horas de elevador por operador.
    
    Args:
        df (DataFrame): DataFrame base processado
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Horas de elevador por operador
    """
    # Agrupar por operador
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Somar horas de elevador da base de cálculo
        horas_elevador_sum = round(dados_op['Horas elevador'].sum(), 2)
        
        resultados.append({
            'Operador': operador,
            'Horas': horas_elevador_sum
        })
    
    return pd.DataFrame(resultados)

def calcular_motor_ocioso_colhedora(base_calculo):
    """
    Calcula o percentual de motor ocioso por operador para colhedoras.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de motor ocioso por operador
    """
    # Agrupar por operador
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Motor Ocioso = Parado Com Motor Ligado / Motor Ligado
        parado_motor_sum = round(dados_op['Parado Com Motor Ligado'].sum(), 4)
        motor_ligado_sum = round(dados_op['Motor Ligado'].sum(), 4)
        
        percentual = calcular_porcentagem(parado_motor_sum, motor_ligado_sum)
        
        resultados.append({
            'Operador': operador,
            'Porcentagem': percentual
        })
    
    return pd.DataFrame(resultados)

def calcular_uso_gps_colhedora(df, base_calculo):
    """
    Calcula o percentual de uso de GPS por operador para colhedoras.
    
    Args:
        df (DataFrame): DataFrame base processado
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de uso de GPS por operador
    """
    # Agrupar por operador
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados base para este operador e grupo
        filtro_base = (df['Operador'] == operador) & (df['Grupo Equipamento/Frente'] == grupo)
        dados_op_base = df[filtro_base]
        
        # Calcular tempo total trabalhando
        tempo_trabalhando = round(dados_op_base[
            (dados_op_base['Estado'].isin(['TRABALHANDO', 'COLHEITA']))
        ]['Diferença_Hora'].sum(), 4)
        
        # Calcular tempo com GPS ativo
        tempo_gps_ativo = round(dados_op_base[
            (dados_op_base['Estado'].isin(['TRABALHANDO', 'COLHEITA'])) &
            (dados_op_base['RTK (Piloto Automatico)'] == 1) &
            (dados_op_base['Velocidade'] > 0)
        ]['Diferença_Hora'].sum(), 4)
        
        # Calcular percentual em formato decimal (0-1)
        percentual = calcular_porcentagem(tempo_gps_ativo, tempo_trabalhando)
        
        # Garantir que não ultrapasse 100% (1.0)
        percentual = min(percentual, 1.0)
        
        resultados.append({
            'Operador': operador,
            'Porcentagem': percentual
        })
    
    return pd.DataFrame(resultados)

# Funções para transbordos
def processar_arquivo_transbordo(caminho_arquivo):
    """
    Processa o arquivo TXT ou CSV e retorna o DataFrame com as transformações necessárias para transbordos.
    
    Args:
        caminho_arquivo (str): Caminho do arquivo TXT ou CSV de entrada
    
    Returns:
        DataFrame: DataFrame processado com todas as transformações
    """
    # Lista de codificações para tentar
    codificacoes = ['utf-8', 'latin1', 'ISO-8859-1', 'cp1252']
    
    for codificacao in codificacoes:
        try:
            # Leitura do arquivo
            df = pd.read_csv(caminho_arquivo, sep=';', encoding=codificacao)
            print(f"Arquivo lido com sucesso usando {codificacao}! Total de linhas: {len(df)}")
            
            # Verificar se o DataFrame está vazio
            if len(df) == 0:
                print(f"O arquivo {caminho_arquivo} contém apenas cabeçalhos sem dados.")
                for col in COLUNAS_DESEJADAS_TRANSBORDOS:
                    if col not in df.columns:
                        df[col] = np.nan
                colunas_existentes = [col for col in COLUNAS_DESEJADAS_TRANSBORDOS if col in df.columns]
                colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS_TRANSBORDOS]
                return df[colunas_existentes + colunas_extras]
            
            # Limpeza de espaços extras nos nomes das colunas
            df.columns = df.columns.str.strip()
            
            # Verificar se 'Data/Hora' existe e processá-la
            if 'Data/Hora' in df.columns:
                df[['Data', 'Hora']] = df['Data/Hora'].str.split(' ', expand=True)
                df = df.drop(columns=['Data/Hora'])
            
            # Conversão e cálculo de diferenças de hora
            if isinstance(df['Hora'].iloc[0], str):
                df['Hora'] = pd.to_datetime(df['Hora'], format='%H:%M:%S', errors='coerce')
            
            # Calcular a diferença de hora se ainda não existir
            if 'Diferença_Hora' not in df.columns or df['Diferença_Hora'].isna().any():
                df['Diferença_Hora'] = df['Hora'].diff().dt.total_seconds() / 3600
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if pd.isna(x) or x < 0 else x)
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if x > 0.50 else round(x, 4))
            else:
                df['Diferença_Hora'] = pd.to_numeric(df['Diferença_Hora'].astype(str).str.strip(), errors='coerce')
                df['Diferença_Hora'] = df['Diferença_Hora'].fillna(0)
                df['Diferença_Hora'] = df['Diferença_Hora'].apply(lambda x: 0 if x > 0.50 else round(x, 4))
            
            # Conversão de Motor Ligado para formato numérico
            if 'Motor Ligado' in df.columns:
                if df['Motor Ligado'].dtype == 'object':
                    df['Motor Ligado'] = df['Motor Ligado'].replace({'LIGADO': 1, 'DESLIGADO': 0})
                df['Motor Ligado'] = pd.to_numeric(df['Motor Ligado'], errors='coerce').fillna(0).astype(int)
            
            # Cálculos específicos para transbordos
            RPM_MINIMO = 300
            
            # Verificar e calcular "Parado Com Motor Ligado" se necessário
            if 'Parado Com Motor Ligado' not in df.columns:
                df['Parado Com Motor Ligado'] = ((df['Velocidade'] == 0) & 
                                               (df['RPM Motor'] >= RPM_MINIMO)).astype(int)
            
            # Verificar se Horas Produtivas já existe
            if 'Horas Produtivas' not in df.columns or df['Horas Produtivas'].isna().any():
                df['Horas Produtivas'] = df.apply(
                    lambda row: round(row['Diferença_Hora'], 4) if row['Grupo Operacao'] == 'Produtiva' else 0,
                    axis=1
                )
            else:
                df['Horas Produtivas'] = pd.to_numeric(df['Horas Produtivas'].astype(str).str.strip(), errors='coerce')
                df['Horas Produtivas'] = df['Horas Produtivas'].fillna(0)
            
            # Coluna de GPS - Para transbordos
            if 'RTK (Piloto Automatico)' in df.columns:
                df['GPS'] = df.apply(
                    lambda row: row['Diferença_Hora'] if row.get('RTK (Piloto Automatico)', 0) == 1 
                    and row['Velocidade'] > 0 and row['Grupo Operacao'] == 'Produtiva' else 0,
                    axis=1
                )
            else:
                df['GPS'] = 0
            
            # Limpeza e organização das colunas
            df = df.drop(columns=COLUNAS_REMOVER_TRANSBORDOS, errors='ignore')
            
            # Garantir que todas as colunas desejadas existam
            for col in COLUNAS_DESEJADAS_TRANSBORDOS:
                if col not in df.columns:
                    df[col] = np.nan
            
            # Reorganizar as colunas na ordem desejada
            colunas_existentes = [col for col in COLUNAS_DESEJADAS_TRANSBORDOS if col in df.columns]
            colunas_extras = [col for col in df.columns if col not in COLUNAS_DESEJADAS_TRANSBORDOS]
            df = df[colunas_existentes + colunas_extras]
            
            return df
            
        except UnicodeDecodeError:
            print(f"Tentativa com codificação {codificacao} falhou, tentando próxima codificação...")
            continue
        except Exception as e:
            print(f"Erro ao processar o arquivo com codificação {codificacao}: {str(e)}")
            continue
    
    print(f"Erro: Não foi possível ler o arquivo {caminho_arquivo} com nenhuma das codificações tentadas.")
    return None

def calcular_base_calculo_transbordo(df):
    """
    Calcula a tabela de Base Calculo a partir do DataFrame processado para transbordos.
    
    Args:
        df (DataFrame): DataFrame processado
    
    Returns:
        DataFrame: Tabela Base Calculo com todas as métricas calculadas
    """
    # Extrair combinações únicas de Equipamento, Grupo Equipamento/Frente e Operador
    combinacoes = df[['Equipamento', 'Grupo Equipamento/Frente', 'Operador']].drop_duplicates().reset_index(drop=True)
    
    # Filtrar operadores excluídos
    combinacoes = combinacoes[~combinacoes['Operador'].isin(OPERADORES_EXCLUIR)]
    
    # Inicializar as colunas de métricas
    resultados = []
    
    # Calcular as métricas para cada combinação
    for idx, row in combinacoes.iterrows():
        equipamento = row['Equipamento']
        grupo = row['Grupo Equipamento/Frente']
        operador = row['Operador']
        
        # Filtrar dados para esta combinação
        filtro = (df['Equipamento'] == equipamento) & \
                (df['Grupo Equipamento/Frente'] == grupo) & \
                (df['Operador'] == operador)
        
        dados_filtrados = df[filtro]
        
        # Horas totais - manter mais casas decimais para cálculos intermediários
        horas_totais = round(dados_filtrados['Diferença_Hora'].sum(), 4)
        
        # Horas Produtivas
        horas_produtivas = round(dados_filtrados['Horas Produtivas'].sum(), 4)
        
        # GPS para transbordos
        gps = round(dados_filtrados['GPS'].sum(), 4)
        
        # % Utilização GPS (em decimal 0-1)
        utilizacao_gps = calcular_porcentagem(gps, horas_produtivas)
        
        # Motor Ligado
        motor_ligado = round(dados_filtrados[dados_filtrados['Motor Ligado'] == 1]['Diferença_Hora'].sum(), 4)
        
        # Parado com Motor Ligado
        parado_motor_ligado = round(dados_filtrados[dados_filtrados['Parado Com Motor Ligado'] == 1]['Diferença_Hora'].sum(), 4)
        
        # % Parado com motor ligado (em decimal 0-1)
        percent_parado_motor = calcular_porcentagem(parado_motor_ligado, motor_ligado)
        
        # Falta de Apontamento - Contabilizar apenas registros explicitamente marcados
        falta_apontamento = round(dados_filtrados[
            (dados_filtrados['Motor Ligado'] == 1) & 
            (
                (dados_filtrados['Codigo da Operacao'] == 8340) |
                (dados_filtrados['Codigo da Operacao'].astype(str).str.startswith('8340')) |
                (dados_filtrados['Operacao'].astype(str).str.contains('FALTA DE APONTAMENTO', case=False))
            )
        ]['Diferença_Hora'].sum(), 4)
        
        # % Falta de apontamento (em decimal 0-1)
        percent_falta_apontamento = calcular_porcentagem(falta_apontamento, motor_ligado)
        
        resultados.append({
            'Equipamento': equipamento,
            'Grupo Equipamento/Frente': grupo,
            'Operador': operador,
            'Horas totais': horas_totais,
            'Horas Produtivas': horas_produtivas,
            'GPS': gps,
            '% Utilização GPS': utilizacao_gps,
            'Motor Ligado': motor_ligado,
            'Parado Com Motor Ligado': parado_motor_ligado,
            '% Parado com motor ligado': percent_parado_motor,
            'Falta de Apontamento': falta_apontamento,
            '% Falta de Apontamento': percent_falta_apontamento
        })
    
    return pd.DataFrame(resultados)

def calcular_eficiencia_energetica_transbordo(base_calculo):
    """
    Calcula a eficiência energética por operador para transbordos.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Eficiência energética por operador
    """
    # Agrupar por operador
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Eficiência Energética para transbordos = Horas Produtivas / Horas Totais
        horas_produtivas_sum = round(dados_op['Horas Produtivas'].sum(), 4)
        horas_totais_sum = round(dados_op['Horas totais'].sum(), 4)
        
        # Calcular eficiência
        eficiencia = calcular_porcentagem(horas_produtivas_sum, horas_totais_sum)
        
        # Garantir que não ultrapasse 100%
        eficiencia = min(eficiencia, 1.0)
        
        resultados.append({
            'Operador': operador,
            'Eficiência': eficiencia
        })
    
    return pd.DataFrame(resultados)

def calcular_motor_ocioso_transbordo(base_calculo):
    """
    Calcula o percentual de motor ocioso por operador para transbordos.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de motor ocioso por operador
    """
    # Agrupar por operador
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Motor Ocioso = Parado Com Motor Ligado / Motor Ligado
        parado_motor_sum = round(dados_op['Parado Com Motor Ligado'].sum(), 4)
        motor_ligado_sum = round(dados_op['Motor Ligado'].sum(), 4)
        
        percentual = calcular_porcentagem(parado_motor_sum, motor_ligado_sum)
        
        resultados.append({
            'Operador': operador,
            'Porcentagem': percentual
        })
    
    return pd.DataFrame(resultados)

def calcular_falta_apontamento(base_calculo):
    """
    Calcula o percentual de falta de apontamento por operador.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de falta de apontamento por operador
    """
    # Agrupar por operador
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Usar os valores já calculados em base_calculo
        falta_apontamento_sum = round(dados_op['Falta de Apontamento'].sum(), 4)
        motor_ligado_sum = round(dados_op['Motor Ligado'].sum(), 4)
        
        # Calcular percentual
        percentual = calcular_porcentagem(falta_apontamento_sum, motor_ligado_sum)
        
        resultados.append({
            'Operador': operador,
            'Porcentagem': percentual
        })
    
    return pd.DataFrame(resultados)

def calcular_uso_gps_transbordo(base_calculo):
    """
    Calcula o percentual de uso de GPS por operador para transbordos.
    
    Args:
        base_calculo (DataFrame): Tabela Base Calculo
    
    Returns:
        DataFrame: Percentual de uso de GPS por operador
    """
    # Agrupar por operador
    operadores = base_calculo[['Operador', 'Grupo Equipamento/Frente']].drop_duplicates()
    resultados = []
    
    for _, row in operadores.iterrows():
        operador = row['Operador']
        grupo = row['Grupo Equipamento/Frente']
        
        # Filtrar dados para este operador e grupo
        filtro = (base_calculo['Operador'] == operador) & (base_calculo['Grupo Equipamento/Frente'] == grupo)
        dados_op = base_calculo[filtro]
        
        # Uso GPS = GPS / Horas Produtivas
        gps_sum = round(dados_op['GPS'].sum(), 4)
        horas_produtivas_sum = round(dados_op['Horas Produtivas'].sum(), 4)
        
        percentual = calcular_porcentagem(gps_sum, horas_produtivas_sum)
        
        # Garantir que não ultrapasse 100% (1.0)
        percentual = min(percentual, 1.0)
        
        resultados.append({
            'Operador': operador,
            'Porcentagem': percentual
        })
    
    return pd.DataFrame(resultados)

def adicionar_planilhas_ao_excel(writer, planilhas_dados, tipo):
    """
    Adiciona um conjunto de planilhas ao arquivo Excel com prefixos apropriados.
    
    Args:
        writer: O objeto ExcelWriter
        planilhas_dados: Um dicionário com nome_planilha -> DataFrame
        tipo: 'CD' para colhedoras ou 'TT' para transbordos
    """
    workbook = writer.book
    
    # Adicionar cada planilha
    for nome_planilha, df in planilhas_dados.items():
        # Adicionar prefixo
        nome_com_prefixo = f"{tipo}_{nome_planilha}"
        
        # Salvar o DataFrame na planilha
        df.to_excel(writer, sheet_name=nome_com_prefixo, index=False)
        
        # Obter a planilha para aplicar formatação
        worksheet = workbook[nome_com_prefixo]
        
        # Aplicar formatação específica com base no tipo de planilha
        if 'Disponibilidade' in nome_planilha:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=2)  # Coluna B (Disponibilidade)
                cell.number_format = '0.00%'
                
        elif 'Eficiência' in nome_planilha:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=2)  # Coluna B (Eficiência)
                cell.number_format = '0.00%'
                
        elif 'Hora Elevador' in nome_planilha:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=2)  # Coluna B (Horas)
                cell.number_format = '0.00'
                
        elif 'Motor Ocioso' in nome_planilha or 'Uso GPS' in nome_planilha or 'Falta de Apontamento' in nome_planilha:
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=2)  # Coluna B (Porcentagem)
                cell.number_format = '0.00%'
                
        elif 'Horas por Frota' in nome_planilha:
            for row in range(2, worksheet.max_row + 1):
                cell_b = worksheet.cell(row=row, column=2)  # Coluna B (Horas Registradas)
                cell_b.number_format = '0.00'
                cell_c = worksheet.cell(row=row, column=3)  # Coluna C (Diferença para 24h)
                cell_c.number_format = '0.00'
                
        elif 'Base Calculo' in nome_planilha:
            # Formatar colunas de Base Calculo conforme o tipo
            if tipo == 'CD':  # Colhedoras
                for row in range(2, worksheet.max_row + 1):
                    # Formatar colunas decimais
                    columns_decimal = [4, 5, 7, 8, 10, 12]
                    for col in columns_decimal:
                        if col <= worksheet.max_column:
                            cell = worksheet.cell(row=row, column=col)
                            cell.number_format = '0.00'
                    
                    # Formatar colunas de porcentagem
                    columns_percent = [6, 9, 11, 13]
                    for col in columns_percent:
                        if col <= worksheet.max_column:
                            cell = worksheet.cell(row=row, column=col)
                            cell.number_format = '0.00%'
            
            else:  # Transbordos
                for row in range(2, worksheet.max_row + 1):
                    # Formatar colunas decimais
                    columns_decimal = [4, 5, 6, 8, 9, 11]
                    for col in columns_decimal:
                        if col <= worksheet.max_column:
                            cell = worksheet.cell(row=row, column=col)
                            cell.number_format = '0.00'
                    
                    # Formatar colunas de porcentagem
                    columns_percent = [7, 10, 12]
                    for col in columns_percent:
                        if col <= worksheet.max_column:
                            cell = worksheet.cell(row=row, column=col)
                            cell.number_format = '0.00%'

def processar_arquivos_unificados():
    """
    Processa todos os arquivos de colhedoras e transbordos, combinando-os em um único arquivo Excel
    com prefixos 'CD_' para colhedoras e 'TT_' para transbordos.
    """
    # Obter o diretório onde está o script
    diretorio_script = os.path.dirname(os.path.abspath(__file__))
    
    # Diretório raiz do projeto
    diretorio_raiz = os.path.dirname(diretorio_script)
    
    # Diretórios para dados de entrada e saída
    diretorio_dados = os.path.join(diretorio_raiz, "dados")
    diretorio_colhedoras = os.path.join(diretorio_raiz, "dados", "colhedoras")
    diretorio_transbordos = os.path.join(diretorio_raiz, "dados", "transbordos")
    diretorio_saida = os.path.join(diretorio_raiz, "output")
    
    # Verificar se os diretórios existem, caso contrário criar
    for diretorio in [diretorio_dados, diretorio_colhedoras, diretorio_transbordos, diretorio_saida]:
        if not os.path.exists(diretorio):
            os.makedirs(diretorio)
    
    # Caminho completo para o arquivo Excel unificado
    caminho_saida_unificado = os.path.join(diretorio_saida, arquivo_saida_unificado)
    
    # Criar o objeto ExcelWriter para o arquivo unificado
    writer = pd.ExcelWriter(caminho_saida_unificado, engine='openpyxl')
    workbook = writer.book
    
    # Lista para controlar quais equipamentos já foram processados
    equipamentos_processados = []
    
    # Processar arquivos de colhedoras
    print("="*80)
    print("Processando arquivos de COLHEDORAS...")
    
    # Lista de diretórios para buscar arquivos de colhedoras
    diretorios_busca_colhedoras = [diretorio_dados, diretorio_colhedoras]
    arquivos_colhedoras = []
    
    for diretorio in diretorios_busca_colhedoras:
        # Adicionar arquivos TXT sempre
        arquivos_colhedoras += glob.glob(os.path.join(diretorio, "RV Colhedora*.txt"))
        arquivos_colhedoras += glob.glob(os.path.join(diretorio, "*colhedora*.txt"))
        arquivos_colhedoras += glob.glob(os.path.join(diretorio, "colhedora*.txt"))
        
        # Adicionar arquivos CSV apenas se processCsv for True
        if processCsv:
            arquivos_colhedoras += glob.glob(os.path.join(diretorio, "RV Colhedora*.csv"))
            arquivos_colhedoras += glob.glob(os.path.join(diretorio, "*colhedora*.csv"))
            arquivos_colhedoras += glob.glob(os.path.join(diretorio, "colhedora*.csv"))
    
    # Filtrar arquivos que contenham "transbordo" no nome (case insensitive)
    arquivos_colhedoras = [arquivo for arquivo in arquivos_colhedoras if "transbordo" not in os.path.basename(arquivo).lower()]
    
    # Remover possíveis duplicatas
    arquivos_colhedoras = list(set(arquivos_colhedoras))
    
    print(f"Encontrados {len(arquivos_colhedoras)} arquivos de colhedoras para processar.")
    
    # Processar cada arquivo de colhedora
    for arquivo in arquivos_colhedoras:
        print(f"\nProcessando arquivo de colhedora: {os.path.basename(arquivo)}")
        
        # Processar o arquivo base
        df_base = processar_arquivo_colhedora(arquivo)
        if df_base is None or len(df_base) == 0:
            print(f"Arquivo {os.path.basename(arquivo)} sem dados válidos. Pulando.")
            continue
        
        # Extrair os equipamentos deste arquivo
        equipamentos = df_base['Equipamento'].unique()
        
        # Verificar se algum equipamento já foi processado e remover dos dados
        equipamentos_a_processar = [eq for eq in equipamentos if eq not in equipamentos_processados]
        
        if not equipamentos_a_processar:
            print(f"Todos os equipamentos de {os.path.basename(arquivo)} já foram processados. Pulando.")
            continue
        
        # Filtrar dados apenas para equipamentos ainda não processados
        df_base_filtrado = df_base[df_base['Equipamento'].isin(equipamentos_a_processar)]
        
        # Calcular a Base Calculo
        base_calculo = calcular_base_calculo_colhedora(df_base_filtrado)
        
        # Calcular as métricas auxiliares
        disp_mecanica = calcular_disponibilidade_mecanica(df_base_filtrado)
        eficiencia_energetica = calcular_eficiencia_energetica_colhedora(base_calculo)
        hora_elevador = calcular_hora_elevador(df_base_filtrado, base_calculo)
        motor_ocioso = calcular_motor_ocioso_colhedora(base_calculo)
        uso_gps = calcular_uso_gps_colhedora(df_base_filtrado, base_calculo)
        horas_por_frota = calcular_horas_por_frota(df_base_filtrado)
        
        # Preparar dicionário de planilhas para adicionar
        planilhas_colhedoras = {
            'BASE': df_base_filtrado,
            'Base Calculo': base_calculo,
            '1_Disponibilidade Mecânica': disp_mecanica,
            '2_Eficiência Energética': eficiencia_energetica,
            '3_Hora Elevador': hora_elevador,
            '4_Motor Ocioso': motor_ocioso,
            '5_Uso GPS': uso_gps,
            'Horas por Frota': horas_por_frota
        }
        
        # Adicionar planilhas ao Excel
        adicionar_planilhas_ao_excel(writer, planilhas_colhedoras, 'CD')
        
        # Marcar estes equipamentos como processados
        equipamentos_processados.extend(equipamentos_a_processar)
    
    # Processar arquivos de transbordos
    print("\n" + "="*80)
    print("Processando arquivos de TRANSBORDOS...")
    
    # Lista de diretórios para buscar arquivos de transbordos
    diretorios_busca_transbordos = [diretorio_dados, diretorio_transbordos]
    arquivos_transbordos = []
    
    for diretorio in diretorios_busca_transbordos:
        # Adicionar arquivos TXT sempre
        arquivos_transbordos += glob.glob(os.path.join(diretorio, "RV Transbordo*.txt"))
        arquivos_transbordos += glob.glob(os.path.join(diretorio, "*transbordo*.txt"))
        arquivos_transbordos += glob.glob(os.path.join(diretorio, "frente*transbordos*.txt"))
        arquivos_transbordos += glob.glob(os.path.join(diretorio, "transbordo*.txt"))
        
        # Adicionar arquivos CSV apenas se processCsv for True
        if processCsv:
            arquivos_transbordos += glob.glob(os.path.join(diretorio, "RV Transbordo*.csv"))
            arquivos_transbordos += glob.glob(os.path.join(diretorio, "*transbordo*.csv"))
            arquivos_transbordos += glob.glob(os.path.join(diretorio, "frente*transbordos*.csv"))
            arquivos_transbordos += glob.glob(os.path.join(diretorio, "transbordo*.csv"))
    
    # Filtrar arquivos que contenham "colhedora" no nome (case insensitive)
    arquivos_transbordos = [arquivo for arquivo in arquivos_transbordos if "colhedora" not in os.path.basename(arquivo).lower()]
    
    # Remover possíveis duplicatas
    arquivos_transbordos = list(set(arquivos_transbordos))
    
    print(f"Encontrados {len(arquivos_transbordos)} arquivos de transbordos para processar.")
    
    # Limpar a lista de equipamentos processados para transbordos
    equipamentos_processados = []
    
    # Processar cada arquivo de transbordo
    for arquivo in arquivos_transbordos:
        print(f"\nProcessando arquivo de transbordo: {os.path.basename(arquivo)}")
        
        # Processar o arquivo base
        df_base = processar_arquivo_transbordo(arquivo)
        if df_base is None or len(df_base) == 0:
            print(f"Arquivo {os.path.basename(arquivo)} sem dados válidos. Pulando.")
            continue
        
        # Extrair os equipamentos deste arquivo
        equipamentos = df_base['Equipamento'].unique()
        
        # Verificar se algum equipamento já foi processado e remover dos dados
        equipamentos_a_processar = [eq for eq in equipamentos if eq not in equipamentos_processados]
        
        if not equipamentos_a_processar:
            print(f"Todos os equipamentos de {os.path.basename(arquivo)} já foram processados. Pulando.")
            continue
        
        # Filtrar dados apenas para equipamentos ainda não processados
        df_base_filtrado = df_base[df_base['Equipamento'].isin(equipamentos_a_processar)]
        
        # Calcular a Base Calculo
        base_calculo = calcular_base_calculo_transbordo(df_base_filtrado)
        
        # Calcular as métricas auxiliares
        disp_mecanica = calcular_disponibilidade_mecanica(df_base_filtrado)
        eficiencia_energetica = calcular_eficiencia_energetica_transbordo(base_calculo)
        motor_ocioso = calcular_motor_ocioso_transbordo(base_calculo)
        falta_apontamento = calcular_falta_apontamento(base_calculo)
        uso_gps = calcular_uso_gps_transbordo(base_calculo)
        horas_por_frota = calcular_horas_por_frota(df_base_filtrado)
        
        # Preparar dicionário de planilhas para adicionar
        planilhas_transbordos = {
            'BASE': df_base_filtrado,
            'Base Calculo': base_calculo,
            '1_Disponibilidade Mecânica': disp_mecanica,
            '2_Eficiência Energética': eficiencia_energetica,
            '3_Motor Ocioso': motor_ocioso,
            '4_Falta de Apontamento': falta_apontamento,
            '5_Uso GPS': uso_gps,
            'Horas por Frota': horas_por_frota
        }
        
        # Adicionar planilhas ao Excel
        adicionar_planilhas_ao_excel(writer, planilhas_transbordos, 'TT')
        
        # Marcar estes equipamentos como processados
        equipamentos_processados.extend(equipamentos_a_processar)
    
    # Salvar o arquivo Excel unificado
    writer.close()
    print("\n" + "="*80)
    print(f"Arquivo unificado gerado com sucesso em: {caminho_saida_unificado}")
    print("="*80)

if __name__ == "__main__":
    print("="*80)
    print("PROCESSAMENTO UNIFICADO DE COLHEDORAS E TRANSBORDOS")
    print(f"Processamento de arquivos CSV: {'ATIVADO' if processCsv else 'DESATIVADO'}")
    print("Gerando um único arquivo Excel contendo todas as planilhas com prefixos:")
    print("  - 'CD_' para dados de colhedoras")
    print("  - 'TT_' para dados de transbordos")
    print("="*80)
    processar_arquivos_unificados()
    print("\nProcessamento concluído!") 